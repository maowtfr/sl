from os import chmod
from altair.vegalite.v4.schema.core import FontStyle
from numpy import True_
import streamlit as st
import PyPDF2
# from bokeh.plotting import figure
import altair as alt
import pandas as pd
#import re
#import shutil
import os
#import time
from pathlib import Path
from shutil import copyfile
from openpyxl import Workbook, load_workbook
from datetime import date
import time

st. set_page_config(layout="wide")


def Replace(str1):
    maketrans = str1.maketrans
    final = str1.translate(maketrans(',.', '.,', ' '))
    return final.replace(',', ", ")


st.markdown(
    '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@4.5.3/dist/css/bootstrap.min.css" integrity="sha384-TX8t27EcRE3e/ihU7zmQxVncDAy5uIKz4rEkgIXeMed4M0jlfIDPvg6uqKI2xXr2" crossorigin="anonymous">',
    unsafe_allow_html=True,
)
query_params = st.experimental_get_query_params()
# , "Hydrogeologie", "Geotechnik","Gebäude-Wiki", "LAGA Boden", "Anregungen und Hilfe"]
tabs = ["LAGA Bauschutt", "LAGA Boden"]#, "LAGA Boden", "Gebäude-Wiki", "Geotechnik", "Hydrogeologie" ]
if "tab" in query_params:
    active_tab = query_params["tab"][0]
else:
    active_tab = "LAGA Bauschutt"

if active_tab not in tabs:
    st.experimental_set_query_params(tab="LAGA Bauschutt")
    active_tab = "LAGA Bauschutt"

li_items = "".join(
    f"""
    <li class="nav-item">
        <a class="nav-link{' active' if t==active_tab else ''}" href="/?tab={t}">{t}</a>
    </li>
    """
    for t in tabs
)
tabs_html = f"""
    <ul class="nav nav-tabs">
    {li_items}
    </ul>
"""


st.markdown(tabs_html, unsafe_allow_html=True)
st.markdown("<br>", unsafe_allow_html=True)


if active_tab == "LAGA Bauschutt":
    # page_bg_img = '''<style>body {background-image: url("https://images.unsplash.com/photo-1542281286-9e0a16bb7366");background-size: cover;}</style>'''

    # st.markdown(page_bg_img, unsafe_allow_html=True)
    st.subheader('Zuordnungswerte Bauschutt LAGA M 20')
    erlass = '[Erlasses des niedersächsischen Umweltministeriums vom 20.12.2011](https://www.ngsmbh.de/bin/pdfs/Erlass_MU_201211.pdf)'
    st.markdown('**Einstellungen**')
    lagatext = '[LAGA M20 TR Boden](https://www.ngsmbh.de/bin/pdfs/Zuordnungswerte.pdf)'
    depV_quelle = '[DepV (2009)](https://www.gesetze-im-internet.de/depv_2009/DepV.pdf)'
    bauschuttquelle2003 = '[LAGA (2003)](https://www.laga-online.de/documents/m20_nov2003u1997_2_1517834540.pdf)'

    excelausgabe = st.checkbox(
        'Ausgabe in xlsx.-Excel-Datei', value=True)
    if excelausgabe == True:
        st.markdown(
            '**Achtung**: Ist diese Option aktiv, wird in das Excel-Blatt zum Stand 2021 geschrieben.')
    elif excelausgabe == False:
        st.markdown(
            '**Achtung**: Die Option ist aktuell deaktiviert. Es wird daher eine alternative .xlsx erzeugt (Prototyp)')

    Recyclingbaustoffcheck_erweitert = st.checkbox(
        'Die Probe wird als Recyclingbaustoff (z.B. Vorabsiebmaterial), d.h. als nicht aufbereiteter Bauschutt, für Rekultivierungszwecke und Geländeauffüllungen in der Einbauklasse 1 verwendet. In Abweichung wird zzgl. die Einbauklasse 2 berücksichtigt.', value=True)
    if Recyclingbaustoffcheck_erweitert:
        st.markdown(
            '**Achtung**: Ist diese Option deaktiviert, wird ausschließlich nach LAGA Bauschutt bewertet.')
    elif Recyclingbaustoffcheck_erweitert == True:
        st.markdown('**Achtung**: Die Option ist aktuell deaktiviert. Die Zuordnung erfolgt ausschließlich nach LAGA Bauschutt und somit ohne Berücksichtigung jedweder Einbauklassen nach LAGA Boden.')
    anwendung_sonderfall_pak16 = st.checkbox(
        'Einzelfallkriterium für erhöhten Grenzwert zur Klassifikation der PAK n. EPA vorhanden', value=False)
    if anwendung_sonderfall_pak16:
        st.markdown(
            '**Achtung**: Betreffend der Bewertungsgrundlagen der LAGA 1997 / 2003 konnte in Einzelfällen von der Klassifikation der jeweiligen Grenzwerte zur Klassifizierung als Z1.1, Z1.2, Z2 sowie >Z2 entsprechend der definierten Vorgaben abgewichen werden. Gemäß der Vorbemerkung zur Veröffentlichung des PDF-Dokumentes der LAGA-Mitteilung 20 "Anforderungen an die stoffliche Verwertung von mineralischen Abfällen" - Technische Regeln vom 05.06.2012 existieren diese alternativen Grenzwerte für Z1.1 und Z1.2 nicht mehr, während der Grenzwert [Z2]->[>Z2] weiterhin bestehen bleibt. Da die Option gegenwärtig aktiviert ist, ist der Grenzwert für die Zuordnung [Z2]->[>Z2] der PAK n. EPA von ursprünglich 75 mg/kg auf 100 mg/kg erhöht.')

    pak_graph = st.checkbox(
        'Ausgabe einer grafischen Übersicht der PAK-Verteilung', value=False)
    if pak_graph == True:
        st.markdown(
            '**Achtung:**: Es wird eine grafische Übersicht der PAK-Verteilung erzeugt. Gegenwärtig ist die Möglichkeit der Auswertung eines PAK-Profilmusters nicht implementiert.')
    # elif pak_graph == False:
    #     st.markdown(
    #         '**Achtung**: Die Option ist aktuell deaktiviert. Es wird daher eine alternative .xlsx erzeugt (Prototyp)')
    pcb_graph = st.checkbox(
        'Ausgabe einer grafischen Übersicht der PCB-Verteilung', value=False)
    if pcb_graph == True:
        st.markdown(
            '**Achtung:**: Es wird eine grafische Übersicht der Verteilung der analysierten PCB erzeugt.')
    depVcheck = st.checkbox(
        'Ausgabe der DepV-Klassifikation der Einzelparameter', value=False)
    if depVcheck:
        st.markdown(
            "**Achtung**. Aktuell ist diese Funktion nicht implementiert!")
    df = pd.read_excel("C:/QLaga0.913 AlphaRelease/user.xlsx",
                       sheet_name='Tabelle1', header=None)  # C:\Users\Frank\Desktop\Programmierung\Aktuelle Projekte\St0706\Streamlit #C:/Users/0z/Desktop/Programmierung/Aktuelle Projekte/St0706/Streamlit/user.xlsx
    user1 = (df.iat[0, 3])
    user2 = (df.iat[1, 3])
    user3 = (df.iat[2, 3])
    user4 = (df.iat[3, 3])
    user5 = (df.iat[4, 3])
    user6 = (df.iat[5, 3])
    user7 = (df.iat[6, 3])
    user8 = (df.iat[7, 3])
    user9 = (df.iat[8, 3])
    user10 = (df.iat[9, 3])
    user11 = (df.iat[10, 3])
    user12 = (df.iat[11, 3])
    user13 = (df.iat[12, 3])
    user14 = (df.iat[13, 3])
    user15 = (df.iat[14, 3])
    user16 = (df.iat[15, 3])
    user17 = (df.iat[16, 3])

    user1_name= (df.iat[0,0])
    user2_name= (df.iat[1,0])
    user3_name= (df.iat[2,0])
    user4_name= (df.iat[3,0])
    user5_name= (df.iat[4,0])
    user6_name= (df.iat[5,0])
    user7_name= (df.iat[6,0])
    user8_name= (df.iat[7,0])
    user9_name= (df.iat[8,0])
    user10_name= (df.iat[9,0])
    user11_name= (df.iat[10,0])
    user12_name= (df.iat[11,0])
    user13_name= (df.iat[12,0])
    user14_name= (df.iat[13,0])
    user15_name= (df.iat[14,0])
    user16_name= (df.iat[15,0])
    user17_name= (df.iat[16,0])

    # my_bar = st.progress(0)
    # for percent_complete in range(10):
    #     time.sleep(0.1)
    #     my_bar.progress(percent_complete + 1)
    option = st.selectbox('Bearbeiter:',
                          (user1, user2, user3, user4, user5, user6, user7, user8, user9, user10, user11, user12, user13, user14, user15, user16, user17))
    if option == user1:
        bearbeiter = user1_name
        telefonnummer = df.iat[0, 2]
        email = df.iat[0, 1]
    elif option == user2:
        bearbeiter = user2_name
        telefonnummer = telefonnummer = df.iat[1, 2]
        email = email = df.iat[1, 1]
    elif option == user3:
        bearbeiter = user3_name
        telefonnummer = telefonnummer = df.iat[2, 2]
        email = email = df.iat[2, 1]
    elif option == user4:
        bearbeiter = user4_name
        telefonnummer = telefonnummer = df.iat[3, 2]
        email = email = df.iat[3, 1]
    elif option == user5:
        bearbeiter = user5_name
        telefonnummer = telefonnummer = df.iat[4, 2]
        email = email = df.iat[4, 1]
    elif option == user6:
        bearbeiter = user6_name
        telefonnummer = telefonnummer = df.iat[5, 2]
        email = email = df.iat[5, 1]
    elif option == user7:
        bearbeiter = user7_name
        telefonnummer = telefonnummer = df.iat[6, 2]
        email = email = df.iat[6, 1]
    elif option == user8:
        bearbeiter = user8_name
        telefonnummer = telefonnummer = df.iat[7, 2]
        email = email = df.iat[7, 1]
    elif option == user9:
        bearbeiter = user9_name
        telefonnummer = telefonnummer = df.iat[8, 2]
        email = email = df.iat[8, 1]
    elif option == user10:
        bearbeiter = user10_name
        telefonnummer = telefonnummer = df.iat[9, 2]
        email = email = df.iat[9, 1]
    elif option == user11:
        bearbeiter = user11_name
        telefonnummer = telefonnummer = df.iat[10, 2]
        email = email = df.iat[10, 1]
    elif option == user12:
        bearbeiter = user12_name
        telefonnummer = telefonnummer = df.iat[11, 2]
        email = email = df.iat[11, 1]
    elif option == user13:
        bearbeiter = user13_name
        telefonnummer = telefonnummer = df.iat[12, 2]
        email = email = df.iat[12, 1]
    elif option == user14:
        bearbeiter = user14_name
        telefonnummer = telefonnummer = df.iat[13, 2]
        email = email = df.iat[13, 1]
    elif option == user15:
        bearbeiter = user15_name
        telefonnummer = telefonnummer = df.iat[14, 2]
        email = email = df.iat[14, 1]
    elif option == user16:
        bearbeiter = user16_name
        telefonnummer = telefonnummer = df.iat[15, 2]
        email = email = df.iat[15, 1]
    elif option == user17:
        bearbeiter = user17_name
        telefonnummer = telefonnummer = df.iat[16, 2]
        email = email = df.iat[16, 1]
    # elif option == user2:
    #     bearbeiter = user2
    #     telefonnummer = telefonnummer = df.iat[1, 2]
    #     email = email = df.iat[1, 1]

    Vorhabenbezeichnung = st.text_input(
        "Vorhabenbezeichnung / Kurzbeschreibung des Projektes", "")
    depV_check_text_1 = 'Die Klassifikation der Zuordnungswerte für Bauschutt vor der Aufbereitung bei unspezifischen Verdacht richtet sich nach dem Mindestuntersuchungsumfang gem. Tab. II 1.4-1 sowie den Zuordnungswerten Feststoff und Eluat für Recyclingbaustoffe \ nicht aufbereiteten Bauschutt gem. Tab. II.1.4-5 ('
    depV_check_text_2 = ', S.76 ff.]. Zusätzlich wird für die Einzelparameter eine Klassifikation gem. '
    Recyclingbaustoff_check_text = 'Die Probe wird nach Tabelle II.1.4-5 gem. '+bauschuttquelle2003 + \
        ' [S.79] als Recyclingbaustoff (z.B. Vorabsiebmaterial), d.h. als nicht aufbereiteter Bauschutt, als Bodenmaterial für Rekultivierungszwecke und Geländeauffüllungen in der Einbauklasse 1 verwendet. Für diese Option ist die Untersuchung von Arsen und Schwermetallen - mit Ausnahme Quecksilber - erforderlich, d.h. es gelten die Kriterien und Zuordnungswerte Z1 (Z 1.1 und Z 1.2) der Technischen Regeln Boden.'
    Bauschutt_text = 'Die Klassifikation für Bauschutt erfolgt nach '
    Projektnummer = st.text_input("Projektnummer")
    st.markdown('**Ausgewählte Option**')

    if depVcheck and not Recyclingbaustoffcheck_erweitert:
        st.info(depV_check_text_1+bauschuttquelle2003 +
                depV_check_text_2+depV_quelle+" ausgegeben.")
    elif Recyclingbaustoffcheck_erweitert and not depVcheck:
        st.info(Recyclingbaustoff_check_text+".")
    elif Recyclingbaustoffcheck_erweitert and depVcheck:
        st.info(Recyclingbaustoff_check_text +
                ' Zusätzlich wird für die Einzelparameter eine Klassifikation der Einzelparameter gem.'+depV_quelle+" ausgegeben.")
    elif not depVcheck and not Recyclingbaustoffcheck_erweitert:
        st.info(depV_check_text_1+bauschuttquelle2003+', S.76 ff.].')
    excelausgabe_manipuliert = 0
    excelausgabe_ = 0
    # with st.beta_expander("Import der PDF-Datei"):
    st.markdown('**Auswahl der zu importierenden PDF-Datei**')
    single_file = st.file_uploader(
        "Bitte nun eine .pdf Datei auswählen, welche die Kriterien nach LAGA Bauschutt erfüllt. Die Datei kann entweder manuell ausgewählt oder aber in das u.a. Feld geschoben werden.")

    if single_file:
        pdfReader = PyPDF2.PdfFileReader(single_file)
        # Auftraggeber
        pageObj = pdfReader.getPage(0)
        pageObj_O = pageObj.extractText()
        auftraggeber_A_raw = pageObj_O.split('Auftraggeber:')[1]
        Auftraggeber = auftraggeber_A_raw.split('Herr')[0]
        # print("ERSTE SEITE*******************************")
        #print(pageObj_O)
        # print("ERSTE SEITE*******************************ENDE")
        # Zeitraum der Prüfung
        pageObj_O = pageObj.extractText()
        zeitraumpruefung_A_raw = pageObj_O.split(
            'Zeitraum der Prüfung:')[1]
        zeitraumpruefung = zeitraumpruefung_A_raw.split('Prüfauftrag')[0]
        # Datum der Bearbeitung
        aktuellesDatum = date.today()

        pageObj = pdfReader.getPage(1)
        pageObj_A = pageObj.extractText()
        Datum_A = pageObj_A.split('Prüfbericht')[0]
        pbn_raw = pageObj_A.split('Probenbezeichnung')[0]
        pbn_raw = pbn_raw.split('Prüfbericht:')[1]
        pbn_raw = pbn_raw.lstrip()  # lstrip entfernt Leerzeichen
        pbn_raw = list(pbn_raw)
        Prüfberichtnummer_A = []  # Prüfberichtnummer
        #print(pageObj_O)
        print(pageObj_A)
        for i in pbn_raw:
            if i.isdigit():
                Prüfberichtnummer_A.append(i)
            else:
                break

        Prüfberichtnummer_A = ''.join(Prüfberichtnummer_A)

        probbez_A = []
        Probenbezeichnung_A = []
        cut_intervall_probbez = len(Prüfberichtnummer_A)
        for i in pbn_raw and range(cut_intervall_probbez, len(pbn_raw)):
            probbez_A.append(pbn_raw[i])
        Probenbezeichnung_A = ''.join(probbez_A)

        #
        entnahmedatum_A_raw = pageObj_A.split('Labornummer:')[1]
        entnahmedatum_A_raw = entnahmedatum_A_raw.split('Verfahren')[0]
        entnahmedatum_A_raw = entnahmedatum_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        entnahmedatum_raw = list(entnahmedatum_A_raw)
        Entnahmedatum_A = []
        Entnahmedatum_A = ''.join(entnahmedatum_A_raw)
        if 'DIN EN 14346' in Entnahmedatum_A:  # Falls "alte PDF Datei gegeben"
            Entnahmedatum_A = Entnahmedatum_A.split('DIN EN 14346')[0]
        # Labornummer
        #
        #
        #
        labornummer_A_raw = pageObj_A.split('Labornummer:')[0]
        labornummer_A_raw = labornummer_A_raw.split('Material:')[1]
        labornummer_A_raw = labornummer_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        labornummer_raw = list(labornummer_A_raw)
        Labornummer_A = []
        Labornummer_A = ''.join(labornummer_A_raw)

        # Material
        #
        #
        #
        material_A_raw = pageObj_A.split('Probenahmedatum:')[1]
        material_A_raw = material_A_raw.split('Material:')[0]
        material_A_raw = material_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        material_raw = list(material_A_raw)
        Material_A = []
        Material_A = ''.join(material_A_raw)
        # Feststoffparameter
        # Trockenrückstand
        #
        trockenrückstand_A_raw = pageObj_A.split('Trockenrückstand')[1]
        trockenrückstand_A_raw = trockenrückstand_A_raw.split('DIN EN ISO 11885')[
            0]
        # lstrip entfernt Leerzeichen
        trockenrückstand_A_raw = trockenrückstand_A_raw.lstrip()
        trockenrückstand_raw = list(trockenrückstand_A_raw)
        Trockenrückstand_A = []
        Trockenrückstand_A = ''.join(trockenrückstand_A_raw)

        udB = []

        if 'Arsen1' in pageObj_A:
            arsen_A_raw = pageObj_A.split('Arsen1')[1]
            Bestimmungsgrenze_Arsen = '1'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Arsen als 1 vor')
        if 'DIN EN ISO 11885' in arsen_A_raw:
            arsen_A_raw = arsen_A_raw.split('DIN EN ISO 11885')[0]
        else:
            arsen_A_raw = arsen_A_raw.split('EN ISO 11885')[0]
        arsen_A_raw = arsen_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        arsen_raw = list(arsen_A_raw)
        Arsen_A = []
        Arsen_A = ''.join(arsen_A_raw)
        if Arsen_A == 'u, d, B, ' or Arsen_A == 'u.d.B.':
            Arsen_A = 0
            udB_as = '  [u.d.B.]'
            udb_conv_as = True
        else:
            Arsen_A = Replace(Arsen_A)
            Arsen_A = float(Arsen_A)
            udB_as = ''
            udb_conv_as = False
        arsen_A_raw_Einheit = 'mg/kg'

        # Blei
        #
        if 'Blei0,2' in pageObj_A:
            blei_A_raw = pageObj_A.split('Blei0,2')[1]
            Bestimmungsgrenze_Blei = '0,2'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Blei als 0,2 vor')
        if 'DIN EN ISO 11885' in blei_A_raw:
            blei_A_raw = blei_A_raw.split('DIN EN ISO 11885')[0]
        else:
            blei_A_raw = blei_A_raw.split('EN ISO 11885')[0]
        blei_A_raw = blei_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        blei_raw = list(blei_A_raw)
        Blei_A = []
        Blei_A = ''.join(blei_A_raw)
        if Blei_A == 'u, d, B, ' or Blei_A == 'u.d.B.':
            Blei_A = 0
            udB_pb = '[u.d.B.]'
            udb_conv_pb = True
        else:
            Blei_A = Replace(Blei_A)
            Blei_A = float(Blei_A)
            udB_pb = ''
            udb_conv_pb = False
        Blei_A_Einheit = 'mg/kg'

        # Cadmium
        #
        if 'Cadmium0,1' in pageObj_A:
            cadmium_A_raw = pageObj_A.split('Cadmium0,1')[1]
            Bestimmungsgrenze_Cadmium = '0,1'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Cadmium als 0,1 vor')
        if 'DIN EN ISO 11885' in cadmium_A_raw:
            cadmium_A_raw = cadmium_A_raw.split('DIN EN ISO 11885')[0]
        else:
            cadmium_A_raw = cadmium_A_raw.split('EN ISO 11885')[0]
        cadmium_A_raw = cadmium_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        cadmium_raw = list(cadmium_A_raw)
        Cadmium_A = []
        Cadmium_A = ''.join(cadmium_A_raw)
        if Cadmium_A == 'u, d, B, ' or Cadmium_A == 'u.d.B.':
            Cadmium_A = 0
            udB_cd = ' [u.d.B.]'
            udb_conv_cd = True
        else:
            Cadmium_A = Replace(Cadmium_A)
            Cadmium_A = float(Cadmium_A)
            udB_cd = ''
            udb_conv_cd = False
        Cadmium_A_Einheit = 'mg/kg'

        # Chrom
        #
        if 'Chrom0,2' in pageObj_A:
            chrom_A_raw = pageObj_A.split('Chrom0,2')[1]
            Bestimmungsgrenze_Chrom = '0,2'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Chrom als 0,2 vor')
        if 'DIN EN ISO 11885' in chrom_A_raw:
            chrom_A_raw = chrom_A_raw.split('DIN EN ISO 11885')[0]
        else:
            chrom_A_raw = chrom_A_raw.split('EN ISO 11885')[0]
        chrom_A_raw = chrom_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        chrom_raw = list(chrom_A_raw)
        Chrom_A = []
        Chrom_A = ''.join(chrom_A_raw)

        Chrom_A_Einheit = 'mg/kg'

        # Kupfer
        #
        if 'Kupfer0,2' in pageObj_A:
            kupfer_A_raw = pageObj_A.split('Kupfer0,2')[1]
            Bestimmungsgrenze_Kupfer = '0,2'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Kupfer als 0,2 vor')
        if 'DIN EN ISO 11885' in kupfer_A_raw:
            kupfer_A_raw = kupfer_A_raw.split('DIN EN ISO 11885')[0]
        else:
            kupfer_A_raw = kupfer_A_raw.split('EN ISO 11885')[0]
        kupfer_A_raw = kupfer_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        kupfer_raw = list(kupfer_A_raw)
        Kupfer_A = []
        Kupfer_A = ''.join(kupfer_A_raw)
        if Chrom_A == 'u, d, B, ' or Chrom_A == 'u.d.B.':
            Chrom_A = 0
            udB_cr = ' [u.d.B.]'
            udb_conv_cr = True
        else:
            Chrom_A = Replace(Chrom_A)
            Chrom_A = float(Chrom_A)
            udB_cr = ''
            udb_conv_cr = False
        if Kupfer_A == 'u, d, B, ' or Kupfer_A == 'u.d.B.':
            Kupfer_A = 0
            udB_cu = ' [u.d.B.]'
            udb_conv_cu = True
        else:
            Kupfer_A = Replace(Kupfer_A)
            Kupfer_A = float(Kupfer_A)
            udB_cu = ''
            udb_conv_cu = False
        Kupfer_A_Einheit = 'mg/kg'

        # Nickel
        #
        if 'Nickel0,5' in pageObj_A:
            nickel_A_raw = pageObj_A.split('Nickel0,5')[1]
            Bestimmungsgrenze_Nickel = '0,5'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Nickel als 0,5 vor')
        if 'mg/kg TS' in nickel_A_raw:
            nickel_A_raw = nickel_A_raw.split('mg/kg TS')[0]
            if 'DIN EN ISO 12846' in nickel_A_raw:
                nickel_A_raw = nickel_A_raw.split('DIN EN ISO 12846')[0]
            else:
                nickel_A_raw = nickel_A_raw.split('EN ISO 12846')[0]
        nickel_A_raw = nickel_A_raw.lstrip()  # lstrip entfernt Leerzeichen

        nickel_raw = list(nickel_A_raw)
        Nickel_A = []
        Nickel_A = ''.join(nickel_A_raw)
        if Nickel_A == 'u, d, B, ' or Nickel_A == 'u.d.B.':
            Nickel_A = 0
            udB_ni = ' [u.d.B.]'
            udb_conv_ni = True
        else:
            Nickel_A = Replace(Nickel_A)
            Nickel_A = float(Nickel_A)
            udB_ni = ''
            udb_conv_ni = False
        Nickel_A_Einheit = 'mg/kg'

        # Quecksilber
        #
        if 'Quecksilber0,1' in pageObj_A:
            quecksilber_A_raw = pageObj_A.split('Quecksilber0,1')[1]
            Bestimmungsgrenze_Quecksilber = '0,1'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Quecksilber als 0,1 vor')
        if 'DIN EN ISO 11885' in quecksilber_A_raw:
            quecksilber_A_raw = quecksilber_A_raw.split(
                'DIN EN ISO 11885')[0]
        else:
            quecksilber_A_raw = quecksilber_A_raw.split('EN ISO 11885')[0]
        quecksilber_A_raw = quecksilber_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        quecksilber_raw = list(quecksilber_A_raw)
        Quecksilber_A = []
        Quecksilber_A = ''.join(quecksilber_A_raw)
        if Quecksilber_A == 'u, d, B, ' or Quecksilber_A == 'u.d.B.':
            Quecksilber_A = 0
            udB_hg = ' [u.d.B.]'
            udb_conv_hg = True
        else:
            Quecksilber_A = Replace(Quecksilber_A)
            Quecksilber_A = float(Quecksilber_A)
            udB_hg = ''
            udb_conv_hg = False
        Quecksilber_A_Einheit = 'mg/kg'

        # Zink
        #
        if 'Zink0,2' in pageObj_A:
            zink_A_raw = pageObj_A.split('Zink0,2')[1]
            Bestimmungsgrenze_Zink = '0,2'
        elif 'Zink0,1' in pageObj_A:
            zink_A_raw = pageObj_A.split('Zink0,1')[1]
            Bestimmungsgrenze_Zink = '0,1'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für Zink vor')
        if 'DIN 38414-17' in zink_A_raw:
            zink_A_raw = zink_A_raw.split('DIN 38414-17')[0]
        else:
            zink_A_raw = zink_A_raw.split('DIN 38414 - S17')[0]
        zink_A_raw = zink_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        zink_raw = list(zink_A_raw)
        Zink_A = []
        Zink_A = ''.join(zink_A_raw)
        if Zink_A == 'u, d, B, ' or Zink_A == 'u.d.B.':
            Zink_A = 0
            udB_zn = ' [u.d.B.]'
            udb_conv_zn = True
        else:
            # print(Zink_A)
            Zink_A = Replace(Zink_A)
            Zink_A = float(Zink_A)
            udB_zn = ''
            udb_conv_zn = False
        Zink_A_Einheit = 'mg/kg'

        # EOX
        #
        if 'EOX0,5' in pageObj_A:
            Eox_A_raw = pageObj_A.split('EOX0,5')[1]
            Bestimmungsgrenze_EOX = '0,5'
        else:
            st.error(
                'Wahrscheinlich liegt ein anderer Grenzwert für EOX als 0,5 vor')
        if 'DIN EN 14039' in Eox_A_raw:
            Eox_A_raw = Eox_A_raw.split('DIN EN 14039')[0]
        else:
            Eox_A_raw = Eox_A_raw.split('DIN 38414 - S17')[0]
        Eox_A_raw = Eox_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Eox_raw = list(Eox_A_raw)
        Eox_A = []
        Eox_A = ''.join(Eox_A_raw)
        if Eox_A == 'u, d, B, ' or Eox_A == 'u.d.B.':
            Eox_A = 0
            udB_eox = ' [u.d.B.]'
            udb_conv_eox = True
        else:
            Eox_A = Replace(Eox_A)
            Eox_A = float(Eox_A)
            udB_eox = ''
            udb_conv_eox = False
        EOX_A_Einheit = 'mg/kg'

        # KW
        if 'Kohlenwasserstoffe\n50' in pageObj_A:
            kw_A_raw_split1 = pageObj_A.split('Kohlenwasserstoffe\n50')[1]
        else:
            kw_A_raw_split1 = pageObj_A.split('Kohlenwasserstoffe50')[1]
        Kw_A_raw = kw_A_raw_split1.split('DIN ISO 18287')[0]
        Kw_A_raw = Kw_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Kw_A_raw = list(Kw_A_raw)
        Kw_A = []
        Kw_A = ''.join(Kw_A_raw)
        if Kw_A == 'u, d, B, ' or Kw_A == 'u.d.B.':
            Kw_A = 0
            udB_kw = ' [u.d.B.]'
            udb_conv_kw = True
        else:
            Kw_A = Replace(Kw_A)
            Kw_A = float(Kw_A)
            udB_kw = ''
            udb_conv_kw = False
        Kw_A_Einheit = 'mg/kg'

        # SummePak16
        if 'Summe der 16 PAK nach EPA\n' in pageObj_A:
            pak16_A_raw_split1 = pageObj_A.split(
                'Summe der 16 PAK nach EPA\n')[1]
        elif 'Summe der 16 PAK nach EPA' in pageObj_A:
            pak16_A_raw_split1 = pageObj_A.split(
                'Summe der 16 PAK nach EPA')[1]
        Pak16_A_raw = pak16_A_raw_split1.split('mg/kg TS')[0]
        Pak16_A_raw = Pak16_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Pak16_A_raw = list(Pak16_A_raw)
        Pak16_A = []
        Pak16_A = ''.join(Pak16_A_raw)
        if Pak16_A == 'u, d, B, ' or Pak16_A == 'u.d.B.':
            Pak16_A = 0
            udB_pak16 = ' [u.d.B.]'
            udb_conv_pak16 = True
        else:
            Pak16_A = Replace(Pak16_A)
            Pak16_A = float(Pak16_A)
            udB_pak16 = ''
            udb_conv_pak16 = False
        # Pak16_A = float(Pak16_A.replace('.', '').replace(',', '.'))
        Pak16_A_Einheit = 'mg/kg'

        # SummePak15
        if 'Summe der 15 PAK ohne Naphthalin' in pageObj_A:
            pak15_A_raw_split1 = pageObj_A.split(
                'Summe der 15 PAK ohne Naphthalin')[1]
        elif 'Summe der 15 PAK (o. Naph.)' in pageObj_A:
            pak15_A_raw_split1 = pageObj_A.split(
                'Summe der 15 PAK (o. Naph.)')[1]
        Pak15_A_raw = pak15_A_raw_split1.split('DIN EN 15308')[0]
        Pak15_A_raw = Pak15_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Pak15_A_raw = list(Pak15_A_raw)
        Pak15_A = []
        Pak15_A = ''.join(Pak15_A_raw)
        if Pak15_A == 'u, d, B, ' or Pak15_A == 'u.d.B.':
            Pak15_A = 0
            udB_pak15 = ' [u.d.B.]'
            udb_conv_pak15 = True
        else:
            Pak15_A = Replace(Pak15_A)
            Pak15_A = float(Pak15_A)
            udB_pak15 = ''
            udb_conv_pak15 = False
        Pak15_A_Einheit = 'mg/kg'

        # SummePCB
        summepcb_A_raw_split1 = pageObj_A.split(
            'Summe der bestimmten PCB')[1]
        Summe_PCB_A_raw = summepcb_A_raw_split1.split('Seite')[0]
        Summe_PCB_A_raw = Summe_PCB_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Summe_PCB_A_raw = list(Summe_PCB_A_raw)
        Summe_PCB_A = []
        Summe_PCB_A = ''.join(Summe_PCB_A_raw)
        Summe_PCB_A_Einheit = 'mg/kg'
        if Summe_PCB_A == 'u, d, B, ' or Summe_PCB_A == 'u.d.B.':
            Summe_PCB_A = 0
            udB_summepcb = ' [u.d.B.]'
            udb_conv_pcb=1
        else:
            Summe_PCB_A = Replace(Summe_PCB_A)
            Summe_PCB_A = float(Summe_PCB_A)
            udB_summepcb = ''
            udb_conv_pcb = False
        # Summe_PCB_A = float(Summe_PCB_A.replace('.', '').replace(',', '.'))
        Summe_PCB_A_Einheit = 'mg/kg'

        # ELUAT
        pageObj_B = pdfReader.getPage(2).extractText()
        # print(pageObj_B)

        # LF
        lf_B_raw_split1 = pageObj_B.split('Elektrische Leitfähigkeit')[1]
        Lf = lf_B_raw_split1.split('DIN EN ISO 10523')[0]
        Lf_B_Einheit = 'µS/cm'

        # pH_carb
        # pH-Wert
        # print(pageObj_B)
        if '(24 h)\n' in pageObj_B:
            # print("24h Umbruch vorhanden")
            phcarb_B_raw_split1 = pageObj_B.split(
                '(24 h)\n')[1]
            # print(phcarb_B_raw_split1)
            pH = phcarb_B_raw_split1.split('DIN')[0]
            pH = pH.replace(",", ".")
            pH_carb = pH
            # pH_24=True
            # print("*PH CARB")
            # print(pH_carb)
        elif '(24 h)' in pageObj_B:
            phcarb_B_raw_split1 = pageObj_B.split(
                '(24 h)')[1]
            pH = phcarb_B_raw_split1.split('DIN EN 27888')[0]
            pH = pH.replace(",", ".")
            # pH24=True
            pH_carb = pH
        else:
            if 'pH-Wert\n' in pageObj_B:
                ph_B_raw_split1 = pageObj_B.split('pH-Wert\n')[1]
            elif 'pH-Wert' in pageObj_B:
                ph_B_raw_split1 = pageObj_B.split('pH-Wert')[1]
            if '\nDIN EN 27888' in pageObj_B:
                pH = ph_B_raw_split1.split('\nDIN EN 27888')[0]
                pH = pH.replace(",", ".")
                # pH=float(pH)
            elif 'DIN EN 27888' in pageObj_B:
                pH = ph_B_raw_split1.split('DIN EN 27888')[0]
                pH = pH.replace(",", ".")
                # pH=float(pH)
            else:
                st.error("Vermutlich existiert kein Wert für den pH")
        # print(pageObj_B)
        if pH:
            pH = pH.lstrip()
            pH = float(pH)
            pH_excel = pH
            pH_24 = False

        if pH_carb:
            pH = pH_carb.lstrip()
            pH = float(pH_carb)
            pH_excel = pH
            pH_24 = True
        # print("******PH 24 h*****")
        # print(pH)
        # print(type(pH))
        # print(pH_carb)
        # else:
        #     st.error('Es gibt ein Problem mit Wertübergabe an pH_carb')
        # Leitfähigkeit
        # print(pageObj_B)
        if 'Elektrische Leitfähigkeit nach Carbonatisierung (24 h)' in pageObj_B:
            # print("******************")
            lf_el__B_raw_split1 = pageObj_B.split(
                'Elektrische Leitfähigkeit nach Carbonatisierung (24 h)')[1]
            # print(lf_el__B_raw_split1)
            Leitfähigkeit = lf_el__B_raw_split1.split('DIN EN ISO 10304-1')[0]
            Leitfähigkeit = Leitfähigkeit.replace(",", ".")
            Leitfähigkeit = float(Leitfähigkeit)
        elif 'Elektrische Leitfähigkeit nach Carbonatisierung (24 h)\n' in pageObj_B:
            # print("******************")
            lf_el__B_raw_split1 = pageObj_B.split(
                'Elektrische Leitfähigkeit nach Carbonatisierung (24 h)\n')[1]
            # print(lf_el__B_raw_split1)
            Leitfähigkeit = lf_el__B_raw_split1.split('DIN EN ISO 10304-1')[0]
            Leitfähigkeit = Leitfähigkeit.replace(",", ".")
            Leitfähigkeit = float(Leitfähigkeit)
        elif 'Elektrische Leitfähigkeit' in pageObj_B:
            lf_el__B_raw_split1 = pageObj_B.split(
                'Elektrische Leitfähigkeit')[1]
            Leitfähigkeit = lf_el__B_raw_split1.split('DIN EN ISO 10304-1')[0]
            Leitfähigkeit = Leitfähigkeit.replace(",", ".")
            Leitfähigkeit = float(Leitfähigkeit)
        elif 'Elektrische Leitfähigkeit' in pageObj_B:
            lf_el__B_raw_split1 = pageObj_B.split(
                'Elektrische Leitfähigkeit')[1]
            Leitfähigkeit = lf_el__B_raw_split1.split('DIN EN ISO 10523')[0]
            Leitfähigkeit = Leitfähigkeit.replace(",", ".")
            Leitfähigkeit = float(Leitfähigkeit)
        else:
            st.error('Es gibt ein Problem bei der Werteübergabe an Leitfähigkeit')
        Leitfähigkeit_Einheit = 'uS/cm'
        # print("*******LFEL******")
        # print(Leitfähigkeit)

        # # # # # # # # # # Chlorid
        # # # # # # # # # chlorid_B_raw_split1 = pageObj_B.split('Chlorid1')[1]
        # # # # # # # # # Cl_raw = chlorid_B_raw_split1.split('DIN EN ISO 10304-1')[0]
        # # # # # # # # # Cl_raw = Cl_raw.lstrip()  # lstrip entfernt Leerzeichen
        # # # # # # # # # Cl_raw = list(Cl_raw)
        # # # # # # # # # Cl = []
        # # # # # # # # # Cl = ''.join(Cl_raw)
        # # # # # # # # # if Cl == 'u, d, B, ' or Cl == 'u.d.B.':
        # # # # # # # # #     Cl = 0
        # # # # # # # # #     udB_cl_el = ' [u.d.B.]'
        # # # # # # # # #     udb_conv_cl=1
        # # # # # # # # # else:
        # # # # # # # # #     Cl = Replace(Cl)
        # # # # # # # # #     Cl = float(Cl)
        # # # # # # # # #     udB_cl_el = ''
        # # # # # # # # # Cl_Einheit = 'mg/l'

        # # # # # # # # # # Sulfat
        # # # # # # # # # sulfat_B_raw_split1 = pageObj_B.split('Sulfat2')[1]
        # # # # # # # # # So4_raw = sulfat_B_raw_split1.split('DI')[0]
        # # # # # # # # # So4_raw = So4_raw.lstrip()  
        # # # # # # # # # So4_raw = list(So4_raw)
        # # # # # # # # # So4 = []
        # # # # # # # # # So4 = ''.join(So4_raw)
        # # # # # # # # # if So4 == 'u, d, B, ' or So4 == 'u.d.B.':
        # # # # # # # # #     So4 = 0
        # # # # # # # # #     udB_sulfat_el = ' [u.d.B.]'
        # # # # # # # # #     udb_conv_so4=1
        # # # # # # # # # else:
        # # # # # # # # #     So4 = Replace(So4)
        # # # # # # # # #     So4 = float(So4)
        # # # # # # # # #     udB_sulfat_el = ''
        # # # # # # # # # So4_Einheit = 'mg/l'

        # Arsen
        arsen_el__B_raw_split1 = pageObj_B.split('Arsen2,5')[1]
        As_el_raw = arsen_el__B_raw_split1.split('DIN EN ISO 17294-2')[0]
        As_el_raw = As_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        As_el_raw = list(As_el_raw)
        As_el = []
        As_el = ''.join(As_el_raw)
        if As_el == 'u, d, B, ' or As_el == 'u.d.B.':
            As_el = 0
            udB_as_el = ' [u.d.B.]'
            udb_conv_as_el=1
        else:
            As_el = Replace(As_el)
            As_el = float(As_el)
            udB_as_el = ''
        As_el_Einheit = 'ug/l'

        # Blei
        blei_el__B_raw_split1 = pageObj_B.split('Blei2,5')[1]
        Pb_el_raw = blei_el__B_raw_split1.split('DIN EN ISO 17294-2')[0]
        Pb_el_raw = Pb_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Pb_el_raw = list(Pb_el_raw)
        Pb_el = []
        Pb_el = ''.join(Pb_el_raw)
        if Pb_el == 'u, d, B, ' or Pb_el == 'u.d.B.':
            Pb_el = 0
            udB_pb_el = ' [u.d.B.]'
            udb_conv_pb_el=1
        else:
            Pb_el = Replace(Pb_el)
            Pb_el = float(Pb_el)
            udB_pb_el = ''
        Pb_el_Einheit = 'ug/l'

        # Cadmium
        cadmium_el__B_raw_split1 = pageObj_B.split('Cadmium0,5')[1]
        Cd_el_raw = cadmium_el__B_raw_split1.split('DIN EN ISO 17294-2')[0]
        Cd_el_raw = Cd_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Cd_el_raw = list(Cd_el_raw)
        Cd_el = []
        Cd_el = ''.join(Cd_el_raw)
        if Cd_el == 'u, d, B, ' or Cd_el == 'u.d.B.':
            Cd_el = 0
            udB_cd_el = ' [u.d.B.]'
            udb_conv_cd_el=1
        else:
            Cd_el = Replace(Cd_el)
            Cd_el = float(Cd_el)
            udB_cd_el = ''
        Cd_el_Einheit = 'ug/l'

        # Chrom
        chrom_el__B_raw_split1 = pageObj_B.split('Chrom5')[1]
        Cr_el_raw = chrom_el__B_raw_split1.split('DIN EN ISO 1')[0]
        Cr_el_raw = Cr_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Cr_el_raw = list(Cr_el_raw)
        Cr_el = []
        Cr_el = ''.join(Cr_el_raw)
        if Cr_el == 'u, d, B, ' or Cr_el == 'u.d.B.':
            Cr_el = 0
            udB_cr_el = ' [u.d.B.]'
            udb_conv_cr_el=1
        else:
            Cr_el = Replace(Cr_el)
            Cr_el = float(Cr_el)
            udB_cr_el = ''
        Cr_el_Einheit = 'ug/l'

        # Kupfer
        kupfer_el__B_raw_split1 = pageObj_B.split('Kupfer10')[1]
        Cu_el_raw = kupfer_el__B_raw_split1.split('DIN EN ISO 17294-2')[0]
        Cu_el_raw = Cu_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Cu_el_raw = list(Cu_el_raw)
        Cu_el = []
        Cu_el = ''.join(Cu_el_raw)
        if Cu_el == 'u, d, B, ' or Cu_el == 'u.d.B.':
            Cu_el = 0
            udB_cu_el = ' [u.d.B.]'
            udb_conv_cu_el=1
        else:
            Cu_el = Replace(Cu_el)
            Cu_el = float(Cu_el)
            udB_cu_el = ''
        Cu_el_Einheit = 'ug/l'

        # Nickel
        nickel_el__B_raw_split1 = pageObj_B.split('Nickel10')[1]
        Ni_el_raw = nickel_el__B_raw_split1.split('DIN EN ISO 12846')[0]
        Ni_el_raw = Ni_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Ni_el_raw = list(Ni_el_raw)
        Ni_el = []
        Ni_el = ''.join(Ni_el_raw)
        if Ni_el == 'u, d, B, ' or Ni_el == 'u.d.B.':
            Ni_el = 0
            udB_ni_el = ' [u.d.B.]'
            udb_conv_ni_el=1
        else:
            Ni_el = Replace(Ni_el)
            Ni_el = float(Ni_el)
            udB_ni_el = ''
        Ni_el_Einheit = 'ug/l'

        # Quecksilber
        quecksilber_el__B_raw_split1 = pageObj_B.split('Quecksilber0,05')[1]
        Hg_el_raw = quecksilber_el__B_raw_split1.split('DIN EN ISO 17294-2')[0]
        Hg_el_raw = Hg_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Hg_el_raw = list(Hg_el_raw)
        Hg_el = []
        Hg_el = ''.join(Hg_el_raw)
        if Hg_el == 'u, d, B, ' or Hg_el == 'u.d.B.':
            Hg_el = 0
            udB_hg_el = ' [u.d.B.]'
            udb_conv_hg_el=1
        else:
            Hg_el = Replace(Hg_el)
            Hg_el = float(Hg_el)
            udB_hg_el = ''
        Hg_el_Einheit = 'ug/l'

        # Zink
        zink_el__B_raw_split1 = pageObj_B.split('Zink10')[1]
        Zn_el_raw = zink_el__B_raw_split1.split('DIN EN ISO 144')[0]
        Zn_el_raw = Zn_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Zn_el_raw = list(Zn_el_raw)
        Zn_el = []
        Zn_el = ''.join(Zn_el_raw)
        if Zn_el == 'u, d, B, ' or Zn_el == 'u.d.B.':
            Zn_el = 0
            udB_zn_el = ' [u.d.B.]'
            udb_conv_zn_el=1
        else:
            Zn_el = Replace(Zn_el)
            Zn_el = float(Zn_el)
            udB_zn_el = ''
        Zn_el_Einheit = 'ug/l'

        # Chlorid
        Cl_el__B_raw_split1 = pageObj_B.split('Chlorid1')[1]
        Cl_el_raw = Cl_el__B_raw_split1.split('DIN EN ISO 10304-1')[0]
        Cl_el_raw = Cl_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        Cl_el_raw = list(Cl_el_raw)
        Cl_el = []
        Cl_el = ''.join(Cl_el_raw)
        # print("******ERSTE INSTANZ CL*****")
        # print(pageObj_B)
        # print(Cl_el)
        if Cl_el == 'u, d, B, ' or Cl_el == 'u.d.B.':
            Cl_el = 0
            udB_cl_el = ' [u.d.B.]'
            udb_conv_cl_el=1
        else:
            Cl_el = Replace(Cl_el)
            Cl_el = float(Cl_el)
            udB_cl_el = ''
        Cl_el_Einheit = 'ug/l'
        # Sulfat
        sulfat_el__B_raw_split1 = pageObj_B.split('Sulfat2')[1]
        So4_el_raw = sulfat_el__B_raw_split1.split('DIN EN ISO 17294-2')[0]
        So4_el_raw = So4_el_raw.lstrip()  # lstrip entfernt Leerzeichen
        So4_el_raw = list(So4_el_raw)
        So4_el = []
        So4_el = ''.join(So4_el_raw)
        # print(pageObj_B)
        # print("******ERSTE INSTANZ SO4*****")
        # print(So4_el)
        if So4_el == 'u, d, B, ' or So4_el == 'u.d.B.':
            So4_el = 0
            udB_sulfat_el = ' [u.d.B.]'
            udb_conv_so4_el=1
        else:
            So4_el = Replace(So4_el)
            So4_el = float(So4_el)
            udB_sulfat_el = ''
        So4_el_Einheit = 'ug/l'
        # Phenolindex
        phenol_B_raw_split1 = pageObj_B.split('Phenolindex0,008')[1]
        Phenol_raw = phenol_B_raw_split1.split('___________')[0]
        Phenol_raw = Phenol_raw.lstrip()  # lstrip entfernt Leerzeichen
        Phenol_raw = list(Phenol_raw)
        Phenol = []
        Phenol = ''.join(Phenol_raw)
        if Phenol == 'u, d, B, ' or Phenol == 'u.d.B.':
            Phenol = 0
            udB_phenol_el = ' [u.d.B.]'
            udb_conv_phenol=1
        else:
            Phenol = Replace(Phenol)
            Phenol = float(Phenol)
            udB_phenol_el = ''
        Phenol_Einheit = 'mg/l'

        # Naphtalin
        naph_A_raw_split1 = pageObj_A.split('Naphthalin0,01')[1]
        Naphthalin_A_raw = naph_A_raw_split1.split('mg/kg TS')[0]
        Naphthalin_A_raw = Naphthalin_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Naphthalin_A_raw = list(Naphthalin_A_raw)
        Naphthalin_A = []
        Naphthalin_A = ''.join(Naphthalin_A_raw)
        if Naphthalin_A == 'u, d, B, ' or Naphthalin_A == 'u.d.B.':
            Naphthalin_A = 0
            udB_naphtalin = ' [u.d.B.]'
        else:
            Naphthalin_A = Replace(Naphthalin_A)
            Naphthalin_A = float(Naphthalin_A)
            udB_naphtalin = ''
        Naphtalin_A_Einheit = 'mg/kg'

        # Acenaphtylen
        acen_A_raw_split1 = pageObj_A.split('Acenaphthylen0,01')[1]
        Acenaphtylen_A_raw = acen_A_raw_split1.split('mg/kg TS')[0]
        Acenaphtylen_A_raw = Acenaphtylen_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Acenaphtylen_A_raw = list(Acenaphtylen_A_raw)
        Acenaphtylen_A = []
        Acenaphtylen_A = ''.join(Acenaphtylen_A_raw)
        if Acenaphtylen_A == 'u, d, B, ' or Acenaphtylen_A == 'u.d.B.':
            Acenaphtylen_A = 0
            udB_acenaphtylen = '[u.d.B.]'
        else:
            Acenaphtylen_A = Replace(Acenaphtylen_A)
            Acenaphtylen_A = float(Acenaphtylen_A)
            udB_acenaphtylen = ''
        Acenaphtylen_A_Einheit = 'mg/kg'

        # Acenaphten
        acenanphten_A_raw_split1 = pageObj_A.split('Acenaphthen0,01')[1]
        Acenaphten_A_raw = acenanphten_A_raw_split1.split(
            'mg/kg TS')[0]
        Acenaphten_A_raw = Acenaphten_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Acenaphten_A_raw = list(Acenaphten_A_raw)
        Acenaphten_A = []
        Acenaphten_A = ''.join(Acenaphten_A_raw)
        if Acenaphten_A == 'u, d, B, ' or Acenaphten_A == 'u.d.B.':
            Acenaphten_A = 0
            udB_acenaphten = '[u.d.B.]'
        else:
            Acenaphten_A = Replace(Acenaphten_A)
            Acenaphten_A = float(Acenaphten_A)
            udB_acenaphten = ''
        Acenaphten_A_Einheit = 'mg/kg'

        # Fluoren
        fluoren_A_raw_split1 = pageObj_A.split('Fluoren0,01')[1]
        Fluoren_A_raw = fluoren_A_raw_split1.split('mg/kg TS')[0]
        Fluoren_A_raw = Fluoren_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Fluoren_A_raw = list(Fluoren_A_raw)
        Fluoren_A = []
        Fluoren_A = ''.join(Fluoren_A_raw)
        if Fluoren_A == 'u, d, B, ' or Fluoren_A == 'u.d.B.':
            Fluoren_A = 0
            udB_fluoren = '[u.d.B.]'
        else:
            Fluoren_A = Replace(Fluoren_A)
            Fluoren_A = float(Fluoren_A)
            udB_fluoren = ''
        Fluoren_A_Einheit = 'mg/kg'

        # Phenanthren
        # print(pageObj_A)
        phenanthren_A_raw_split1 = pageObj_A.split('Phenanthren0,01')[1]
        phenanthren_A_raw = phenanthren_A_raw_split1.split(
            'mg/kg TS')[0]
        Phenanthren_A_raw = phenanthren_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Phenanthren_A_raw = list(Phenanthren_A_raw)
        Phenanthren_A = []
        Phenanthren_A = ''.join(Phenanthren_A_raw)
        if Phenanthren_A == 'u, d, B, ' or Phenanthren_A == 'u.d.B.':
            Phenanthren_A = 0
            udB_Phenanthren = '[u.d.B.]'
        else:
            Phenanthren_A = Replace(Phenanthren_A)
            Phenanthren_A = float(Phenanthren_A)
            udB_Phenanthren = ''
        Phenanthren_A_Einheit = 'mg/kg'

        # Anthracen
        anthracen_A_raw_split1 = pageObj_A.split('Anthracen0,01')[1]
        Anthracen_A_raw = anthracen_A_raw_split1.split(
            'mg/kg TS')[0]
        Anthracen_A_raw = Anthracen_A_raw.lstrip()
        Anthracen_A_raw = list(Anthracen_A_raw)
        Anthracen_A = []
        Anthracen_A = ''.join(Anthracen_A_raw)
        if Anthracen_A == 'u, d, B, ' or Anthracen_A == 'u.d.B.':
            Anthracen_A = 0
            udB_Anthracen = '[u.d.B.]'
        else:
            Anthracen_A = Replace(Anthracen_A)
            Anthracen_A = float(Anthracen_A)
            udB_Anthracen = ''
        Anthracen_A_Einheit = 'mg/kg'

        # Fluoranthen
        fluoranthen_A_raw_split1 = pageObj_A.split('Fluoranthen0,01')[1]
        Fluoranthen_A_raw = fluoranthen_A_raw_split1.split(
            'mg/kg TS')[0]
        Fluoranthen_A_raw = Fluoranthen_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Fluoranthen_A_raw = list(Fluoranthen_A_raw)
        Fluoranthen_A = []
        Fluoranthen_A = ''.join(Fluoranthen_A_raw)
        if Fluoranthen_A == 'u, d, B, ' or Fluoranthen_A == 'u.d.B.':
            Fluoranthen_A = 0
            udB_Fluoranthen = '[u.d.B.]'
        else:
            Fluoranthen_A = Replace(Fluoranthen_A)
            Fluoranthen_A = float(Fluoranthen_A)
            udB_Fluoranthen = ''
        Fluoranthen_A_Einheit = 'mg/kg'

        # Pyren
        pyren_A_raw_split1 = pageObj_A.split('Pyren0,01')[1]
        Pyren_A_raw = pyren_A_raw_split1.split('mg/kg TS')[0]
        Pyren_A_raw = Pyren_A_raw.lstrip()
        Pyren_A_raw = list(Pyren_A_raw)
        Pyren_A = []
        Pyren_A = ''.join(Pyren_A_raw)
        if Pyren_A == 'u, d, B, ' or Pyren_A == 'u.d.B.':
            Pyren_A = 0
            udB_Pyren = '[u.d.B.]'
        else:
            Pyren_A = Replace(Pyren_A)
            Pyren_A = float(Pyren_A)
            udB_Pyren = ''
        Pyren_A_Einheit = 'mg/kg'

        # Benz(a)anthracen
        benzaanthracen_A_raw_split1 = pageObj_A.split('Benz(a)anthracen0,01')[
            1]
        Benzaanthracen_A_raw = benzaanthracen_A_raw_split1.split(
            'mg/kg TS')[0]
        Benzanthracen_A_raw = Benzaanthracen_A_raw.lstrip()
        Benzaanthracen_A_raw = list(Benzaanthracen_A_raw)
        Benzaanthracen_A = []
        Benzaanthracen_A = ''.join(Benzaanthracen_A_raw)
        if Benzaanthracen_A == 'u, d, B, ' or Benzaanthracen_A == 'u.d.B.':
            Benzaanthracen_A = 0
            udB_Benzaanthracen = '[u.d.B.]'
        else:
            Benzaanthracen_A = Replace(Benzaanthracen_A)
            Benzaanthracen_A = float(Benzaanthracen_A)
            udB_Benzaanthracen = ''
        Benzaanthracen_A_Einheit = 'mg/kg'

        # Chrysen
        chrysen_A_raw_split1 = pageObj_A.split('Chrysen0,01')[1]
        Chrysen_A_raw = chrysen_A_raw_split1.split('mg/kg TS')[0]
        Chrysen_A_raw = Chrysen_A_raw.lstrip()
        Chrysen_A_raw = list(Chrysen_A_raw)
        Chrysen_A = []
        Chrysen_A = ''.join(Chrysen_A_raw)
        if Chrysen_A == 'u, d, B, ' or Chrysen_A == 'u.d.B.':
            Chrysen_A = 0
            udB_Chrysen = '[u.d.B.]'
        else:
            Chrysen_A = Replace(Chrysen_A)
            Chrysen_A = float(Chrysen_A)
            udB_Chrysen = ''
        Chrysen_A_Einheit = 'mg/kg'

        # Benzo(b)fluoranthen
        benzobfluoranthen_A_raw_split1 = pageObj_A.split(
            'Benzo(b)fluoranthen0,01')[1]
        Benzobfluoranthen_A_raw = benzobfluoranthen_A_raw_split1.split(
            'mg/kg TS')[0]
        Benzobfluoranthen_A_raw = Benzobfluoranthen_A_raw.lstrip()
        Benzobfluoranthen_A_raw = list(Benzobfluoranthen_A_raw)
        Benzobfluoranthen_A = []
        Benzobfluoranthen_A = ''.join(Benzobfluoranthen_A_raw)
        if Benzobfluoranthen_A == 'u, d, B, ' or Benzobfluoranthen_A == 'u.d.B.':
            Benzobfluoranthen_A = 0
            udB_Benzobfluoranthen = '[u.d.B.]'
        else:
            Benzobfluoranthen_A = Replace(Benzobfluoranthen_A)
            Benzobfluoranthen_A = float(Benzobfluoranthen_A)
            udB_Benzobfluoranthen = ''
        Benzobfluoranthen_A_Einheit = 'mg/kg'

        # Benzo(k)fluoranthen
        benzokfluoranthen_A_raw_split1 = pageObj_A.split(
            'Benzo(k)fluoranthen0,01')[1]
        Benzokfluoranthen_A_raw = benzokfluoranthen_A_raw_split1.split(
            'mg/kg TS')[0]
        Benzokfluoranthen_A_raw = Benzokfluoranthen_A_raw.lstrip()
        Benzokfluoranthen_A_raw = list(Benzokfluoranthen_A_raw)
        Benzokfluoranthen_A = []
        Benzokfluoranthen_A = ''.join(Benzokfluoranthen_A_raw)
        if Benzokfluoranthen_A == 'u, d, B, ' or Benzokfluoranthen_A == 'u.d.B.':
            Benzokfluoranthen_A = 0
            udB_Benzokfluoranthen = '[u.d.B.]'
        else:
            Benzokfluoranthen_A = Replace(Benzokfluoranthen_A)
            Benzokfluoranthen_A = float(Benzokfluoranthen_A)
            udB_Benzokfluoranthen = ''
        Benzokfluoranthen_A_Einheit = 'mg/kg'

        # Benzo(a)pyren
        benzoapyren_A_raw_split1 = pageObj_A.split('Benzo(a)pyren0,01')[1]
        Benzoapyren_A_raw = benzoapyren_A_raw_split1.split(
            'mg/kg TS')[0]
        Benzoapyren_A_raw = Benzoapyren_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Benzoapyren_A_raw = list(Benzoapyren_A_raw)
        Benzoapyren_A = []
        Benzoapyren_A = ''.join(Benzoapyren_A_raw)
        if Benzoapyren_A == 'u, d, B, ' or Benzoapyren_A == 'u.d.B.':
            Benzoapyren_A = 0
            udB_Benzoapyren = '[u.d.B.]'
        else:
            Benzoapyren_A = Replace(Benzoapyren_A)
            Benzoapyren_A = float(Benzoapyren_A)
            udB_Benzoapyren = ''
        Benzoapyren_A_Einheit = 'mg/kg'

        # Indeno(123-cd)pyren
        indeno123cdpyren_A_raw_split1 = pageObj_A.split(
            'Indeno(123-cd)pyren0,01')[1]
        Indeno123cdpyren_A_raw = indeno123cdpyren_A_raw_split1.split(
            'mg/kg TS')[0]
        Indeno123cdpyren_A_raw = Indeno123cdpyren_A_raw.lstrip()
        Indeno123cdpyren_A_raw = list(Indeno123cdpyren_A_raw)
        Indeno123cdpyren_A = []
        Indeno123cdpyren_A = ''.join(Indeno123cdpyren_A_raw)
        if Indeno123cdpyren_A == 'u, d, B, ' or Indeno123cdpyren_A == 'u.d.B.':
            Indeno123cdpyren_A = 0
            udB_Indeno123cdpyren = '[u.d.B.]'
        else:
            Indeno123cdpyren_A = Replace(Indeno123cdpyren_A)
            Indeno123cdpyren_A = float(Indeno123cdpyren_A)
            udB_Indeno123cdpyren = ''
        Indeno123cdpyren_A_Einheit = 'mg/kg'

        # Dibenz(ah)anthracen
        dibenzahanthracen_A_raw_split1 = pageObj_A.split(
            'Dibenz(ah)anthracen0,01')[1]
        Dibenzahanthracen_A_raw = dibenzahanthracen_A_raw_split1.split(
            'mg/kg TS')[0]
        Dibenzahanthracen_A_raw = Dibenzahanthracen_A_raw.lstrip()
        Dibenzahanthracen_A_raw = list(Dibenzahanthracen_A_raw)
        Dibenzahanthracen_A = []
        Dibenzahanthracen_A = ''.join(Dibenzahanthracen_A_raw)
        if Dibenzahanthracen_A == 'u, d, B, ' or Dibenzahanthracen_A == 'u.d.B.':
            Dibenzahanthracen_A = 0
            udB_Dibenzanthracen = '[u.d.B.]'
        else:
            Dibenzahanthracen_A = Replace(Dibenzahanthracen_A)
            Dibenzahanthracen_A = float(Dibenzahanthracen_A)
            udB_Dibenzanthracen = ''
        Dibenzahanthracen_A_Einheit = 'mg/kg'

        # Benzo(ghi)perylen
        benzoghiperylen_A_raw_split1 = pageObj_A.split('Benzo(ghi)perylen0,01')[
            1]
        Benzoghiperylen_A_raw = benzoghiperylen_A_raw_split1.split(
            'mg/kg TS')[0]
        Benzoghiperylen_A_raw = Benzoghiperylen_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Benzoghiperylen_A_raw = list(Benzoghiperylen_A_raw)
        Benzoghiperylen_A = []
        Benzoghiperylen_A = ''.join(Benzoghiperylen_A_raw)
        if Benzoghiperylen_A == 'u, d, B, ' or Benzoghiperylen_A == 'u.d.B.':
            Benzoghiperylen_A = 0
            udB_Benzoghiperylen = '[u.d.B.]'
        else:
            Benzoghiperylen_A = Replace(Benzoghiperylen_A)
            Benzoghiperylen_A = float(Benzoghiperylen_A)
            udB_Benzoghiperylen = ''
        Benzoghiperylen_A_Einheit = 'mg/kg'

        # PCB28
        pcb28_A_raw_split1 = pageObj_A.split('PCB Nr. 280,005')[1]
        Pcb28_A_raw = pcb28_A_raw_split1.split('mg/kg TS')[0]
        Pcb28_A_raw = Pcb28_A_raw.lstrip()
        Pcb28_A_raw = list(Pcb28_A_raw)
        Pcb28_A = []
        Pcb28_A = ''.join(Pcb28_A_raw)
        if Pcb28_A == 'u, d, B, ' or Pcb28_A == 'u.d.B.':
            Pcb28_A = 0
            udBpcb28 = '[u.d.B.]'
        else:
            Pcb28_A = Replace(Pcb28_A)
            Pcb28_A = float(Pcb28_A)
            udBpcb28 = ''
        Pcb28_A_Einheit = 'mg/kg'
        # Pcb28_A = Pcb28_A.replace('.', '').replace(',', '.')

        # PCB52
        pcb52_A_raw_split1 = pageObj_A.split('PCB Nr. 520,005')[1]
        Pcb52_A_raw = pcb52_A_raw_split1.split('mg/kg TS')[0]
        Pcb52_A_raw = Pcb52_A_raw.lstrip()
        Pcb52_A_raw = list(Pcb52_A_raw)
        Pcb52_A = []
        Pcb52_A = ''.join(Pcb52_A_raw)
        if Pcb52_A == 'u, d, B, ' or Pcb52_A == 'u.d.B.':
            Pcb52_A = 0
            udBpcb52 = '[u.d.B.]'
        else:
            Pcb52_A = Replace(Pcb52_A)
            Pcb52_A = float(Pcb52_A)
            udBpcb52 = ''
        Pcb52_A_Einheit = 'mg/kg'

        # PCB101
        pcb101_A_raw_split1 = pageObj_A.split('PCB Nr. 1010,005')[1]
        Pcb101_A_raw = pcb101_A_raw_split1.split('mg/kg TS')[0]
        Pcb101_A_raw = Pcb101_A_raw.lstrip()
        Pcb101_A_raw = list(Pcb101_A_raw)
        Pcb101_A = []
        Pcb101_A = ''.join(Pcb101_A_raw)
        if Pcb101_A == 'u, d, B, ' or Pcb101_A == 'u.d.B.':
            Pcb101_A = 0
            udBpcb101 = '[u.d.B.]'
        else:
            Pcb101_A = Replace(Pcb101_A)
            Pcb101_A = float(Pcb101_A)
            udBpcb101 = ''
        Pcb101_A_Einheit = 'mg/kg'

        # PCB153
        pcb153_A_raw_split1 = pageObj_A.split('PCB Nr. 1530,005')[1]
        Pcb153_A_raw = pcb153_A_raw_split1.split('mg/kg TS')[0]
        Pcb153_A_raw = Pcb153_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Pcb153_A_raw = list(Pcb153_A_raw)
        Pcb153_A = []
        Pcb153_A = ''.join(Pcb153_A_raw)
        if Pcb153_A == 'u, d, B, ' or Pcb153_A == 'u.d.B.':
            Pcb153_A = 0
            udBpcb153 = '[u.d.B.]'
        else:
            Pcb153_A = Replace(Pcb153_A)
            Pcb153_A = float(Pcb153_A)
            udB153 = ''
        Pcb153_A_Einheit = 'mg/kg'

        # PCB138
        pcb138_A_raw_split1 = pageObj_A.split('PCB Nr. 1380,005')[1]
        Pcb138_A_raw = pcb138_A_raw_split1.split('mg/kg TS')[0]
        Pcb138_A_raw = Pcb138_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Pcb138_A_raw = list(Pcb138_A_raw)
        Pcb138_A = []
        Pcb138_A = ''.join(Pcb138_A_raw)
        if Pcb138_A == 'u, d, B, ' or Pcb138_A == 'u.d.B.':
            Pcb138_A = 0
            udBpcb138 = '[u.d.B.]'
        else:
            Pcb138_A = Replace(Pcb138_A)
            Pcb138_A = float(Pcb138_A)
            udBpcb138 = ''
        Pcb138_A_Einheit = 'mg/kg'

        # PCB180
        pcb180_A_raw_split1 = pageObj_A.split('PCB Nr. 1800,005')[1]
        Pcb180_A_raw = pcb180_A_raw_split1.split('mg/kg TS')[0]
        Pcb180_A_raw = Pcb180_A_raw.lstrip()  # lstrip entfernt Leerzeichen
        Pcb180_A_raw = list(Pcb180_A_raw)
        Pcb180_A = []
        Pcb180_A = ''.join(Pcb180_A_raw)
        # Pcb180_A = float(Pcb180_A.replace('.', '').replace(',', '.'))
        if Pcb180_A == 'u, d, B, ' or Pcb180_A == 'u.d.B.':
            Pcb180_A = 0
            udBpcb180 = '[u.d.B.]'
        else:
            Pcb180_A = Replace(Pcb180_A)
            Pcb180_A = float(Pcb180_A)
            udBpcb180 = ''
        Pcb180_A_Einheit = 'mg/kg'
    # Auswertung
        if single_file:
            # ~~~~~~~~~~~~Auswertung~~~~~~~~~~~~~~~~~~~~~
            z_sm_A = []         # Liste für Zuordnungswerte Schwermetalle
            depV_sm = []      # Liste DepV-Werte
            g_abfall = []     # Parameter, für welche die Deklaration "gefährlicher ABfall" besteht
            g_abf = []        # ~als Volltext

            #Arsen_A = 21
            As_Z0_A = 20
            As_Z1_1_A = 30
            As_Z1_2_A = 50
            As_Z2_A = 150

            #Blei_A = 101
            Pb_Z0_A = 100
            Pb_Z1_1_A = 200
            Pb_Z1_2_A = 300
            Pb_Z2_A = 1000

            #Cadmium_A = 0
            Cd_Z0_A = 0.6
            Cd_Z1_1_A = 1
            Cd_Z1_2_A = 3
            Cd_Z2_A = 10

            #Chrom_A = 0
            Cr_Z0_A = 50
            Cr_Z1_1_A = 100
            Cr_Z1_2_A = 200
            Cr_Z2_A = 600

            #Kupfer_A = 0
            Cu_Z0_A = 40
            Cu_Z1_1_A = 100
            Cu_Z1_2_A = 200
            Cu_Z2_A = 600

            #Nickel_A = 0
            Ni_Z0_A = 40
            Ni_Z1_1_A = 100
            Ni_Z1_2_A = 200
            Ni_Z2_A = 600

            #Quecksilber_A = 0
            Hg_Z0_A = 0.3
            Hg_Z1_1_A = 1
            Hg_Z1_2_A = 3
            Hg_Z2_A = 10

            #Zink_A = 0
            Zn_Z0_A = 120
            Zn_Z1_1_A = 300
            Zn_Z1_2_A = 500
            Zn_Z2_A = 1500

            # EOX
            #Eox_A = 0
            Eox_Z0_A = 1
            Eox_Z1_1_A = 3
            Eox_Z1_2_A = 10
            Eox_Z2_A = 15
            Eox_Z0_A_f = 1
            Eox_Z1_1_A_f = 3
            Eox_Z1_2_A_f = 5
            Eox_Z2_A_f = 10

            #Kw_A = 0
            Kw_Z0_A = 100
            Kw_Z1_1_A = 300
            Kw_Z1_2_A = 500
            Kw_Z2_A = 1000
            Kw_Z0_A_f = 100
            Kw_Z1_1_A_f = 300
            Kw_Z1_2_A_f = 500
            Kw_Z2_A_f = 1000

            # PAK16
            Pak16_A = 90
            Pak16_Z0_A = 1
            Pak16_Z1_1_A = 5
            Pak16_Z1_2_A = 15
            
            Pak16_Z0_A_f = 1
            Pak16_Z1_1_A_f = 5
            Pak16_Z1_2_A_f = 15
            if anwendung_sonderfall_pak16:
                Pak16_Z2_A_f = 100
                Pak16_Z2_A = 100
            else:
                Pak16_Z2_A_f = 75
                Pak16_Z2_A = 75

            #Summe_PCB_A = 0.0
            Summe_PCB_Z0_A = 0.02
            Summe_PCB_Z1_1_A = 0.1
            Summe_PCB_Z1_2_A = 0.5
            Summe_PCB_Z2_A = 1
            Summe_PCB_Z0_A_f = 0.02
            Summe_PCB_Z1_1_A_f = 0.1
            Summe_PCB_Z1_2_A_f = 0.5
            Summe_PCB_Z2_A_f = 1

            # pH
            # pH=0
            #pH = 13.1

            # LF
            # Leitfähigkeit = 600
            Leitfähigkeit_Z0_A_el_f = 500
            Leitfähigkeit_Z1_1_A_el_f = 1500
            Leitfähigkeit_Z1_2_A_el_f = 2500
            Leitfähigkeit_Z2_A_el_f = 3000
            Leitfähigkeit_Z0_A_el = 500
            Leitfähigkeit_Z1_1_A_el = 1500
            Leitfähigkeit_Z1_2_A_el = 2500
            Leitfähigkeit_Z2_A_el = 3000

            # Chlorid
            #Cl_el = 0
            Cl_Z0_A_el_f = 10
            Cl_Z1_1_A_el_f = 20
            Cl_Z1_2_A_el_f = 40
            Cl_Z2_A_el_f = 150
            Cl_Z0_A_el = 10
            Cl_Z1_1_A_el = 20
            Cl_Z1_2_A_el = 40
            Cl_Z2_A_el = 150

            # Sulfat
            #So4_el = 0
            So4_Z0_A_el_f = 50
            So4_Z1_1_A_el_f = 150
            So4_Z1_2_A_el_f = 300
            So4_Z2_A_el_f = 600
            So4_Z0_A_el = 50
            So4_Z1_1_A_el = 150
            So4_Z1_2_A_el = 300
            So4_Z2_A_el = 600

            # Arsen_el S.25 Laga 1997

            #As_el = 0
            As_Z0_A_el_f = 10
            As_Z1_1_A_el_f = 10
            As_Z1_2_A_el_f = 40
            As_Z2_A_el_f = 50
            As_Z0_A_el = 10
            As_Z1_1_A_el = 10
            As_Z1_2_A_el = 40
            As_Z2_A_el = 50

            # Blei_el
            #Pb_el = 0
            Pb_Z0_A_el_f = 20
            Pb_Z1_1_A_el_f = 40
            Pb_Z1_2_A_el_f = 100
            Pb_Z2_A_el_f = 100
            Pb_Z0_A_el = 20
            Pb_Z1_1_A_el = 40
            Pb_Z1_2_A_el = 100
            Pb_Z2_A_el = 100

            # Cadmium_el
            #Cd_el = 0
            Cd_Z0_A_el_f = 2
            Cd_Z1_1_A_el_f = 2
            Cd_Z1_2_A_el_f = 5
            Cd_Z2_A_el_f = 5
            Cd_Z0_A_el = 2
            Cd_Z1_1_A_el = 2
            Cd_Z1_2_A_el = 5
            Cd_Z2_A_el = 5

            # Chrom_el
            #Cr_el = 0
            Cr_Z0_A_el_f = 15
            Cr_Z1_1_A_el_f = 30
            Cr_Z1_2_A_el_f = 75
            Cr_Z2_A_el_f = 100
            Cr_Z0_A_el = 15
            Cr_Z1_1_A_el = 30
            Cr_Z1_2_A_el = 75
            Cr_Z2_A_el = 100

            # Kupfer_el
            #Cu_el = 0
            Cu_Z0_A_el_f = 50
            Cu_Z1_1_A_el_f = 50
            Cu_Z1_2_A_el_f = 150
            Cu_Z2_A_el_f = 200
            Cu_Z0_A_el = 50
            Cu_Z1_1_A_el = 50
            Cu_Z1_2_A_el = 150
            Cu_Z2_A_el = 200

            # Nickel_el
            #Ni_el = 0
            Ni_Z0_A_el_f = 40
            Ni_Z1_1_A_el_f = 50
            Ni_Z1_2_A_el_f = 100
            Ni_Z2_A_el_f = 100
            Ni_Z0_A_el = 40
            Ni_Z1_1_A_el = 50
            Ni_Z1_2_A_el = 100
            Ni_Z2_A_el = 100

            # Quecksilber_el
            #Hg_el = 0
            Hg_Z0_A_el_f = 0.2
            Hg_Z1_1_A_el_f = 0.2
            Hg_Z1_2_A_el_f = 1
            Hg_Z2_A_el_f = 2
            Hg_Z0_A_el = 0.2
            Hg_Z1_1_A_el = 0.2
            Hg_Z1_2_A_el = 1
            Hg_Z2_A_el = 2

            # ZInk_el
            Zn_el = 500
            Zn_Z0_A_el_f = 100
            Zn_Z1_1_A_el_f = 100
            Zn_Z1_2_A_el_f = 300
            Zn_Z2_A_el_f = 400
            Zn_Z0_A_el = 100
            Zn_Z1_1_A_el = 100
            Zn_Z1_2_A_el = 300
            Zn_Z2_A_el = 400

            # Phenolindex_el
            #Phenol = 0
            Phenol_Z0_A_el_f = 10
            Phenol_Z1_1_A_el_f = 10
            Phenol_Z1_2_A_el_f = 50
            Phenol_Z2_A_el_f = 100
            Phenol_Z0_A_el = 10
            Phenol_Z1_1_A_el = 10
            Phenol_Z1_2_A_el = 50
            Phenol_Z2_A_el = 100
            # Klassifikation im Feststoff

            Z_prior = []
            O_Grenze_Z = []
            Z_next = []

            if Arsen_A <= As_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_as_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(As_Z0_A)
                Arsen_A_erhöhung = round((((Arsen_A/As_Z0_A)*100)-100), 2)
            if Arsen_A > As_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_as_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(As_Z0_A)
                Arsen_A_erhöhung = round((((Arsen_A/As_Z0_A)*100)-100), 2)
            elif Arsen_A <= As_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_as_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(As_Z0_A)
                Arsen_A_erhöhung = round((((Arsen_A/As_Z0_A)*100)-100), 2)
            elif Arsen_A > As_Z0_A and Arsen_A <= As_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_as_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(As_Z1_1_A)
                Arsen_A_erhöhung = round((((Arsen_A/As_Z0_A)*100)-100), 2)
            elif Arsen_A > As_Z1_1_A and Arsen_A <= As_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_as_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(As_Z1_2_A)
                Arsen_A_erhöhung = round((((Arsen_A/As_Z1_1_A)*100)-100), 2)
            elif Arsen_A > As_Z1_2_A and Arsen_A <= As_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_as_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(As_Z2_A)
                Arsen_A_erhöhung = round((((Arsen_A/As_Z1_2_A)*100)-100), 2)
            elif Arsen_A > As_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_as_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(As_Z2_A)
                Arsen_A_erhöhung = round((((Arsen_A/As_Z2_A)*100)-100), 2)
                depV = True

            if Blei_A <= Pb_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_pb_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Pb_Z0_A)
                Blei_A_erhöhung = round((((Blei_A/Pb_Z0_A)*100)-100), 2)
            if Blei_A > Pb_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_pb_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Pb_Z0_A)
                Blei_A_erhöhung = round((((Blei_A/Pb_Z0_A)*100)-100), 2)
            elif Blei_A <= Pb_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_pb_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Pb_Z0_A)
                Blei_A_erhöhung = round((((Blei_A/Pb_Z0_A)*100)-100), 2)
            elif Blei_A > Pb_Z0_A and Blei_A <= Pb_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_pb_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Pb_Z1_1_A)
                Blei_A_erhöhung = round((((Blei_A/Pb_Z0_A)*100)-100), 2)
            elif Blei_A > Pb_Z1_1_A and Blei_A <= Pb_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pb_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Pb_Z1_2_A)
                Blei_A_erhöhung = round((((Blei_A/Pb_Z1_1_A)*100)-100), 2)
            elif Blei_A > Pb_Z1_2_A and Blei_A <= Pb_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pb_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Pb_Z2_A)
                Blei_A_erhöhung = round((((Blei_A/Pb_Z1_2_A)*100)-100), 2)
            elif Blei_A > Pb_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pb_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Pb_Z2_A)
                Blei_A_erhöhung = round((((Blei_A/Pb_Z2_A)*100)-100), 2)
                depV = True

            if Cadmium_A <= Cd_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_cd_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Cd_Z0_A)
                Cadmium_A_erhöhung = round((((Cadmium_A/Cd_Z0_A)*100)-100), 2)
            if Cadmium_A > Cd_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_cd_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Cd_Z0_A)
                Cadmium_A_erhöhung = round((((Cadmium_A/Cd_Z0_A)*100)-100), 2)
            elif Cadmium_A <= Cd_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_cd_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Cd_Z0_A)
                Cadmium_A_erhöhung = round((((Cadmium_A/Cd_Z0_A)*100)-100), 2)
            elif Cadmium_A > Cd_Z0_A and Cadmium_A <= Cd_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_cd_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cd_Z1_1_A)
                Cadmium_A_erhöhung = round((((Cadmium_A/Cd_Z0_A)*100)-100), 2)
            elif Cadmium_A > Cd_Z1_1_A and Cadmium_A <= Cd_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cd_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cd_Z1_2_A)
                Cadmium_A_erhöhung = round(
                    (((Cadmium_A/Cd_Z1_1_A)*100)-100), 2)
            elif Cadmium_A > Cd_Z1_2_A and Cadmium_A <= Cd_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cd_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cd_Z2_A)
                Cadmium_A_erhöhung = round(
                    (((Cadmium_A/Cd_Z1_2_A)*100)-100), 2)
            elif Cadmium_A > Cd_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cd_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cd_Z2_A)
                Cadmium_A_erhöhung = round((((Cadmium_A/Cd_Z2_A)*100)-100), 2)
                depV = True

            if Chrom_A <= Cr_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_cr_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Cr_Z0_A)
                Chrom_A_erhöhung = round((((Chrom_A/Cr_Z0_A)*100)-100), 2)
            if Chrom_A > Cr_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_cr_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Cr_Z0_A)
                Chrom_A_erhöhung = round((((Chrom_A/Cr_Z0_A)*100)-100), 2)
            elif Chrom_A <= Cr_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_cr_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Cr_Z0_A)
                Chrom_A_erhöhung = round((((Chrom_A/Cr_Z0_A)*100)-100), 2)
            elif Chrom_A > Cr_Z0_A and Chrom_A <= Cr_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_cr_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cr_Z1_1_A)
                Chrom_A_erhöhung = round((((Chrom_A/Cr_Z0_A)*100)-100), 2)
            elif Chrom_A > Cr_Z1_1_A and Chrom_A <= Cr_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cr_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cr_Z1_2_A)
                Chrom_A_erhöhung = round((((Chrom_A/Cr_Z1_1_A)*100)-100), 2)
            elif Chrom_A > Cr_Z1_2_A and Chrom_A <= Cr_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cr_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cr_Z2_A)
                Chrom_A_erhöhung = round((((Chrom_A/Cr_Z1_2_A)*100)-100), 2)
            elif Chrom_A > Cr_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cr_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cr_Z2_A)
                Chrom_A_erhöhung = round((((Chrom_A/Cr_Z2_A)*100)-100), 2)
                depV = True

            if Kupfer_A <= Cu_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_cu_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Cu_Z0_A)
                Kupfer_A_erhöhung = round((((Kupfer_A/Cu_Z0_A)*100)-100), 2)
            if Kupfer_A > Cu_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_cu_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Cu_Z0_A)
                Kupfer_A_erhöhung = round((((Kupfer_A/Cu_Z0_A)*100)-100), 2)
            elif Kupfer_A <= Cu_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_cu_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Cu_Z0_A)
                Kupfer_A_erhöhung = round((((Kupfer_A/Cu_Z0_A)*100)-100), 2)
            elif Kupfer_A > Cu_Z0_A and Kupfer_A <= Cu_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_cu_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cu_Z1_1_A)
                Kupfer_A_erhöhung = round((((Kupfer_A/Cu_Z0_A)*100)-100), 2)
            elif Kupfer_A > Cu_Z1_1_A and Kupfer_A <= Cu_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cu_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cu_Z1_2_A)
                Kupfer_A_erhöhung = round((((Kupfer_A/Cu_Z1_1_A)*100)-100), 2)
            elif Kupfer_A > Cu_Z1_2_A and Kupfer_A <= Cu_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cu_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cu_Z2_A)
                Kupfer_A_erhöhung = round((((Kupfer_A/Cu_Z1_2_A)*100)-100), 2)
            elif Kupfer_A > Cu_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_cu_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cu_Z2_A)
                Kupfer_A_erhöhung = round((((Kupfer_A/Cu_Z2_A)*100)-100), 2)
                depV = True

            if Nickel_A <= Ni_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_ni_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Ni_Z0_A)
                Nickel_A_erhöhung = round((((Nickel_A/Ni_Z0_A)*100)-100), 2)
            if Nickel_A > Ni_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_ni_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Ni_Z0_A)
                Nickel_A_erhöhung = round((((Nickel_A/Ni_Z0_A)*100)-100), 2)
            elif Nickel_A <= Ni_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_ni_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Ni_Z0_A)
                Nickel_A_erhöhung = round((((Nickel_A/Ni_Z0_A)*100)-100), 2)
            elif Nickel_A > Ni_Z0_A and Nickel_A <= Ni_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_ni_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Ni_Z1_1_A)
                Nickel_A_erhöhung = round((((Nickel_A/Ni_Z0_A)*100)-100), 2)
            elif Nickel_A > Ni_Z1_1_A and Nickel_A <= Ni_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_ni_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Ni_Z1_2_A)
                Nickel_A_erhöhung = round((((Nickel_A/Ni_Z1_1_A)*100)-100), 2)
            elif Nickel_A > Ni_Z1_2_A and Nickel_A <= Ni_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_ni_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Ni_Z2_A)
                Nickel_A_erhöhung = round((((Nickel_A/Ni_Z1_2_A)*100)-100), 2)
            elif Nickel_A > Ni_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_ni_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Ni_Z2_A)
                Nickel_A_erhöhung = round((((Nickel_A/Ni_Z2_A)*100)-100), 2)
                depV = True

            if Quecksilber_A <= Hg_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_hg_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Hg_Z0_A)
                Quecksilber_A_erhöhung = round(
                    (((Quecksilber_A/Hg_Z0_A)*100)-100), 2)
            if Quecksilber_A > Hg_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_hg_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Hg_Z0_A)
                Quecksilber_A_erhöhung = round(
                    (((Quecksilber_A/Hg_Z0_A)*100)-100), 2)
            elif Quecksilber_A <= Hg_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_hg_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Hg_Z0_A)
                Quecksilber_A_erhöhung = round(
                    (((Quecksilber_A/Hg_Z0_A)*100)-100), 2)
            elif Quecksilber_A > Hg_Z0_A and Quecksilber_A <= Hg_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_hg_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Hg_Z1_1_A)
                Quecksilber_A_erhöhung = round(
                    (((Quecksilber_A/Hg_Z0_A)*100)-100), 2)
            elif Quecksilber_A > Hg_Z1_1_A and Quecksilber_A <= Hg_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_hg_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Hg_Z1_2_A)
                Quecksilber_A_erhöhung = round(
                    (((Quecksilber_A/Hg_Z1_1_A)*100)-100), 2)
            elif Quecksilber_A > Hg_Z1_2_A and Quecksilber_A <= Hg_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_hg_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Hg_Z2_A)
                Quecksilber_A_erhöhung = round(
                    (((Quecksilber_A/Hg_Z1_2_A)*100)-100), 2)
            elif Quecksilber_A > Hg_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_hg_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Hg_Z2_A)
                Quecksilber_A_erhöhung = round(
                    (((Quecksilber_A/Hg_Z2_A)*100)-100), 2)
                depV = True

            if Zink_A <= Zn_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_zn_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Zn_Z0_A)
                Zink_A_erhöhung = round((((Zink_A/Zn_Z0_A)*100)-100), 2)
            if Zink_A > Zn_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_zn_A = 9
                z_sm_A.append(9)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Zn_Z0_A)
                Zink_A_erhöhung = round((((Zink_A/Zn_Z0_A)*100)-100), 2)
            elif Zink_A <= Zn_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_zn_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Zn_Z0_A)
                Zink_A_erhöhung = round((((Zink_A/Zn_Z0_A)*100)-100), 2)
            elif Zink_A > Zn_Z0_A and Zink_A <= Zn_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_zn_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Zn_Z1_1_A)
                Zink_A_erhöhung = round((((Zink_A/Zn_Z0_A)*100)-100), 2)
            elif Zink_A > Zn_Z1_1_A and Zink_A <= Zn_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_zn_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Zn_Z1_2_A)
                Zink_A_erhöhung = round((((Zink_A/Zn_Z1_1_A)*100)-100), 2)
            elif Zink_A > Zn_Z1_2_A and Zink_A <= Zn_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_zn_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Zn_Z2_A)
                Zink_A_erhöhung = round((((Zink_A/Zn_Z1_2_A)*100)-100), 2)
            elif Zink_A > Zn_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_zn_A = 3  # fehler! bug!!!!!BUG
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Zn_Z2_A)
                Zink_A_erhöhung = round((((Zink_A/Zn_Z2_A)*100)-100), 2)
                depV = True

            if Eox_A <= Eox_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_eox_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Eox_Z0_A)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z0_A)*100)-100), 2)
            if Eox_A > Eox_Z0_A and Eox_A <= Eox_Z1_1_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_eox_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Eox_Z1_1_A_f)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z0_A_f)*100)-100), 2)
            elif Eox_A > Eox_Z1_1_A_f and Eox_A <= Eox_Z1_2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_eox_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Eox_Z1_2_A_f)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z1_1_A)*100)-100), 2)
            elif Eox_A > Eox_Z1_2_A_f and Eox_A <= Eox_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_eox_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Eox_Z2_A_f)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z1_2_A_f)*100)-100), 2)
            elif Eox_A > Eox_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_eox_A = 3  # fehler! bug!!!!!
                z_sm_A.append(3)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Eox_Z2_A_f)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z2_A_f)*100)-100), 2)
                depV = True
            elif Eox_A <= Eox_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_eox_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Eox_Z0_A)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z0_A)*100)-100), 2)
            elif Eox_A > Eox_Z0_A and Eox_A <= Eox_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_eox_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Eox_Z1_1_A)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z0_A)*100)-100), 2)
            elif Eox_A > Eox_Z1_1_A and Eox_A <= Eox_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_eox_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Eox_Z1_2_A)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z1_1_A)*100)-100), 2)
            elif Eox_A > Eox_Z1_2_A and Eox_A <= Eox_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_eox_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Eox_Z2_A)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z1_2_A)*100)-100), 2)
            elif Eox_A > Eox_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_eox_A = 3  # fehler! bug!!!!!
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Eox_Z2_A)
                Eox_A_erhöhung = round((((Eox_A/Eox_Z2_A)*100)-100), 2)
                depV = True

            if Kw_A <= Kw_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_kw_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Kw_Z0_A)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z0_A)*100)-100), 2)
            if Kw_A > Kw_Z0_A and Kw_A <= Kw_Z1_1_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_kw_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Kw_Z1_1_A_f)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z0_A_f)*100)-100), 2)
            elif Kw_A > Kw_Z1_1_A_f and Kw_A <= Kw_Z1_2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_kw_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Kw_Z1_2_A_f)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z1_1_A)*100)-100), 2)
            elif Kw_A > Kw_Z1_2_A_f and Kw_A <= Kw_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_kw_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Kw_Z2_A_f)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z1_2_A_f)*100)-100), 2)
            elif Kw_A > Kw_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_kw_A = 3  # fehler! bug!!!!!
                z_sm_A.append(3)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Kw_Z2_A_f)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z2_A_f)*100)-100), 2)
                # depV = True
            elif Kw_A <= Kw_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_kw_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Kw_Z0_A)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z0_A)*100)-100), 2)
            elif Kw_A > Kw_Z0_A and Kw_A <= Kw_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_kw_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Kw_Z1_1_A)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z0_A)*100)-100), 2)
            elif Kw_A > Kw_Z1_1_A and Kw_A <= Kw_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_kw_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Kw_Z1_2_A)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z1_1_A)*100)-100), 2)
            elif Kw_A > Kw_Z1_2_A and Kw_A <= Kw_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_kw_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Kw_Z2_A)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z1_2_A)*100)-100), 2)
            elif Kw_A > Kw_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_kw_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Kw_Z2_A)
                Kw_A_erhöhung = round((((Kw_A/Kw_Z2_A)*100)-100), 2)
                depV = True

            if Pak16_A <= Pak16_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_pak16_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Pak16_Z0_A)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z0_A)*100)-100), 2)
            if Pak16_A > Pak16_Z0_A and Pak16_A <= Pak16_Z1_1_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pak16_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Pak16_Z1_1_A_f)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z0_A_f)*100)-100), 2)
            elif Pak16_A > Pak16_Z1_1_A_f and Pak16_A <= Pak16_Z1_2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pak16_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Pak16_Z1_2_A_f)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z1_1_A)*100)-100), 2)
            elif Pak16_A > Pak16_Z1_2_A_f and Pak16_A <= Pak16_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pak16_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Pak16_Z2_A_f)
                Pak16_A_erhöhung = round(
                    (((Pak16_A/Pak16_Z1_2_A_f)*100)-100), 2)
            elif Pak16_A > Pak16_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pak16_A = 3  # fehler! bug!!!!!
                z_sm_A.append(3)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Pak16_Z2_A_f)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z2_A_f)*100)-100), 2)
                # depV = True
            elif Pak16_A <= Pak16_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_pak16_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Pak16_Z0_A)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z0_A)*100)-100), 2)
            elif Pak16_A > Pak16_Z0_A and Pak16_A <= Pak16_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_pak16_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Pak16_Z1_1_A)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z0_A)*100)-100), 2)
            elif Pak16_A > Pak16_Z1_1_A and Pak16_A <= Pak16_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pak16_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Pak16_Z1_2_A)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z1_1_A)*100)-100), 2)
            elif Pak16_A > Pak16_Z1_2_A and Pak16_A <= Pak16_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pak16_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Pak16_Z2_A)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z1_2_A)*100)-100), 2)
            elif Pak16_A > Pak16_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pak16_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Pak16_Z2_A)
                Pak16_A_erhöhung = round((((Pak16_A/Pak16_Z2_A)*100)-100), 2)
                depV = True

            if Summe_PCB_A <= Summe_PCB_Z0_A and Recyclingbaustoffcheck_erweitert == False:
                z_pcb_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z0]')
                O_Grenze_Z.append(Summe_PCB_Z0_A)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z0_A)*100)-100), 2)
            if Summe_PCB_A > Summe_PCB_Z0_A and Summe_PCB_A <= Summe_PCB_Z1_1_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pcb_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Summe_PCB_Z1_1_A_f)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z0_A_f)*100)-100), 2)
            elif Summe_PCB_A > Summe_PCB_Z1_1_A_f and Summe_PCB_A <= Summe_PCB_Z1_2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pcb_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Summe_PCB_Z1_2_A_f)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z1_1_A)*100)-100), 2)
            elif Summe_PCB_A > Summe_PCB_Z1_2_A_f and Summe_PCB_A <= Summe_PCB_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pcb_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Summe_PCB_Z2_A_f)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z1_2_A_f)*100)-100), 2)
            elif Summe_PCB_A > Summe_PCB_Z2_A_f and Recyclingbaustoffcheck_erweitert == False:
                z_pcb_A = 3  # fehler! bug!!!!!
                z_sm_A.append(3)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Summe_PCB_Z2_A_f)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z2_A_f)*100)-100), 2)
                # depV = True
            elif Summe_PCB_A <= Summe_PCB_Z0_A and Recyclingbaustoffcheck_erweitert == True:
                z_pcb_A = 0
                z_sm_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Summe_PCB_Z0_A)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z0_A)*100)-100), 2)
            elif Summe_PCB_A > Summe_PCB_Z0_A and Summe_PCB_A <= Summe_PCB_Z1_1_A and Recyclingbaustoffcheck_erweitert == True:
                z_pcb_A = 1.1
                z_sm_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Summe_PCB_Z1_1_A)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z0_A)*100)-100), 2)
            elif Summe_PCB_A > Summe_PCB_Z1_1_A and Summe_PCB_A <= Summe_PCB_Z1_2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pcb_A = 1.2
                z_sm_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Summe_PCB_Z1_2_A)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z1_1_A)*100)-100), 2)
            elif Summe_PCB_A > Summe_PCB_Z1_2_A and Summe_PCB_A <= Summe_PCB_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pcb_A = 2
                z_sm_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Summe_PCB_Z2_A)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z1_2_A)*100)-100), 2)
            elif Summe_PCB_A > Summe_PCB_Z2_A and Recyclingbaustoffcheck_erweitert == True:
                z_pcb_A = 3
                z_sm_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Summe_PCB_Z2_A)
                Summe_PCB_A_erhöhung = round(
                    (((Summe_PCB_A/Summe_PCB_Z2_A)*100)-100), 2)
                depV = True

            z_el_A = []
            # Klassifikation im Eluat
            if As_el <= As_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_as_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(As_Z0_A_el_f)
                As_el_erhöhung = round((((As_el/As_Z0_A_el_f)*100)-100), 2)
            elif As_el > As_Z0_A_el_f and As_el <= As_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_as_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(As_Z1_1_A_el_f)
                As_el_erhöhung = round((((As_el/As_Z0_A_el_f)*100)-100), 2)
            elif As_el > As_Z1_1_A_el_f and As_el <= As_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_as_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(As_Z1_2_A_el_f)
                As_el_erhöhung = round((((As_el/As_Z1_1_A_el_f)*100)-100), 2)
            elif As_el > As_Z1_2_A_el_f and As_el <= As_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_as_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(As_Z2_A_el_f)
                As_el_erhöhung = round((((As_el/As_Z1_2_A_el_f)*100)-100), 2)
            elif As_el > As_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_as_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(As_Z2_A_el_f)
                As_el_erhöhung = round((((As_el/As_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Pb_el <= Pb_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Pb_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Pb_Z0_A_el_f)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z0_A_el_f)*100)-100), 2)
            elif Pb_el > Pb_Z0_A_el_f and Pb_el <= Pb_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Pb_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Pb_Z1_1_A_el_f)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z0_A_el_f)*100)-100), 2)
            elif Pb_el > Pb_Z1_1_A_el_f and Pb_el <= Pb_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Pb_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Pb_Z1_2_A_el_f)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z1_1_A_el_f)*100)-100), 2)
            elif Pb_el > Pb_Z1_2_A_el_f and Pb_el <= Pb_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Pb_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Pb_Z2_A_el_f)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z1_2_A_el_f)*100)-100), 2)
            elif Pb_el > Pb_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Pb_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Pb_Z2_A_el_f)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Cd_el <= Cd_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cd_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Cd_Z0_A_el_f)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z0_A_el_f)*100)-100), 2)
            elif Cd_el > Cd_Z0_A_el_f and Cd_el <= Cd_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cd_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cd_Z1_1_A_el_f)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z0_A_el_f)*100)-100), 2)
            elif Cd_el > Cd_Z1_1_A_el_f and Cd_el <= Cd_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cd_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cd_Z1_2_A_el_f)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z1_1_A_el_f)*100)-100), 2)
            elif Cd_el > Cd_Z1_2_A_el_f and Cd_el <= Cd_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cd_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cd_Z2_A_el_f)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z1_2_A_el_f)*100)-100), 2)
            elif Cd_el > Cd_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cd_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cd_Z2_A_el_f)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Cr_el <= Cr_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cr_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Cr_Z0_A_el_f)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z0_A_el_f)*100)-100), 2)
            elif Cr_el > Cr_Z0_A_el_f and Cr_el <= Cr_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cr_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cr_Z1_1_A_el_f)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z0_A_el_f)*100)-100), 2)
            elif Cr_el > Cr_Z1_1_A_el_f and Cr_el <= Cr_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cr_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cr_Z1_2_A_el_f)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z1_1_A_el_f)*100)-100), 2)
            elif Cr_el > Cr_Z1_2_A_el_f and Cr_el <= Cr_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cr_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cr_Z2_A_el_f)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z1_2_A_el_f)*100)-100), 2)
            elif Cr_el > Cr_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cr_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cr_Z2_A_el_f)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Cu_el <= Cu_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cu_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Cu_Z0_A_el_f)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z0_A_el_f)*100)-100), 2)
            elif Cu_el > Cu_Z0_A_el_f and Cu_el <= Cu_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cu_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cu_Z1_1_A_el_f)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z0_A_el_f)*100)-100), 2)
            elif Cu_el > Cu_Z1_1_A_el_f and Cu_el <= Cu_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cu_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cu_Z1_2_A_el_f)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z1_1_A_el_f)*100)-100), 2)
            elif Cu_el > Cu_Z1_2_A_el_f and Cu_el <= Cu_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cu_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cu_Z2_A_el_f)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z1_2_A_el_f)*100)-100), 2)
            elif Cu_el > Cu_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cu_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cu_Z2_A_el_f)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Ni_el <= Ni_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Ni_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Ni_Z0_A_el_f)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z0_A_el_f)*100)-100), 2)
            elif Ni_el > Ni_Z0_A_el_f and Ni_el <= Ni_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Ni_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Ni_Z1_1_A_el_f)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z0_A_el_f)*100)-100), 2)
            elif Ni_el > Ni_Z1_1_A_el_f and Ni_el <= Ni_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Ni_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Ni_Z1_2_A_el_f)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z1_1_A_el_f)*100)-100), 2)
            elif Ni_el > Ni_Z1_2_A_el_f and Ni_el <= Ni_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Ni_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Ni_Z2_A_el_f)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z1_2_A_el_f)*100)-100), 2)
            elif Ni_el > Ni_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Ni_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Ni_Z2_A_el_f)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Hg_el <= Hg_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Hg_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Hg_Z0_A_el_f)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z0_A_el_f)*100)-100), 2)
            elif Hg_el > Hg_Z0_A_el_f and Hg_el <= Hg_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Hg_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Hg_Z1_1_A_el_f)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z0_A_el_f)*100)-100), 2)
            elif Hg_el > Hg_Z1_1_A_el_f and Hg_el <= Hg_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Hg_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Hg_Z1_2_A_el_f)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z1_1_A_el_f)*100)-100), 2)
            elif Hg_el > Hg_Z1_2_A_el_f and Hg_el <= Hg_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Hg_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Hg_Z2_A_el_f)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z1_2_A_el_f)*100)-100), 2)
            elif Hg_el > Hg_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Hg_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Hg_Z2_A_el_f)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Zn_el <= Zn_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Zn_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Zn_Z0_A_el_f)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z0_A_el_f)*100)-100), 2)
            elif Zn_el > Zn_Z0_A_el_f and Zn_el <= Zn_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Zn_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Zn_Z1_1_A_el_f)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z0_A_el_f)*100)-100), 2)
            elif Zn_el > Zn_Z1_1_A_el_f and Zn_el <= Zn_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Zn_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Zn_Z1_2_A_el_f)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z1_1_A_el_f)*100)-100), 2)
            elif Zn_el > Zn_Z1_2_A_el_f and Zn_el <= Zn_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Zn_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Zn_Z2_A_el_f)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z1_2_A_el_f)*100)-100), 2)
            elif Zn_el > Zn_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Zn_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Zn_Z2_A_el_f)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Cl_el <= Cl_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cl_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Cl_Z0_A_el_f)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z0_A_el_f)*100)-100), 2)
            elif Cl_el > Cl_Z0_A_el_f and Cl_el <= Cl_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cl_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cl_Z1_1_A_el_f)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z0_A_el_f)*100)-100), 2)
            elif Cl_el > Cl_Z1_1_A_el_f and Cl_el <= Cl_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cl_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cl_Z1_2_A_el_f)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z1_1_A_el_f)*100)-100), 2)
            elif Cl_el > Cl_Z1_2_A_el_f and Cl_el <= Cl_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cl_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cl_Z2_A_el_f)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z1_2_A_el_f)*100)-100), 2)
            elif Cl_el > Cl_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Cl_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cl_Z2_A_el_f)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z2_A_el_f)*100)-100), 2)
                depV = True

            if So4_el <= So4_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_So4_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(So4_Z0_A_el_f)
                So4_el_erhöhung = round((((So4_el/So4_Z0_A_el_f)*100)-100), 2)
            elif So4_el > So4_Z0_A_el_f and So4_el <= So4_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_So4_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(So4_Z1_1_A_el_f)
                So4_el_erhöhung = round((((So4_el/So4_Z0_A_el_f)*100)-100), 2)
            elif So4_el > So4_Z1_1_A_el_f and So4_el <= So4_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_So4_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(So4_Z1_2_A_el_f)
                So4_el_erhöhung = round(
                    (((So4_el/So4_Z1_1_A_el_f)*100)-100), 2)
            elif So4_el > So4_Z1_2_A_el_f and So4_el <= So4_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_So4_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(So4_Z2_A_el_f)
                So4_el_erhöhung = round(
                    (((So4_el/So4_Z1_2_A_el_f)*100)-100), 2)
            elif So4_el > So4_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_So4_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(So4_Z2_A_el_f)
                So4_el_erhöhung = round((((So4_el/So4_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Phenol < Phenol_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Phenol_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Phenol_Z0_A_el_f)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z0_A_el_f)*100)-100), 2)
            elif Phenol == Phenol_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Phenol_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Phenol_Z1_1_A_el_f)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z0_A_el_f)*100)-100), 2)
            elif Phenol > Phenol_Z1_1_A_el_f and Phenol <= Phenol_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Phenol_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Phenol_Z1_2_A_el_f)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z1_1_A_el_f)*100)-100), 2)
            elif Phenol > Phenol_Z1_2_A_el_f and Phenol <= Phenol_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Phenol_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Phenol_Z2_A_el_f)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z1_2_A_el_f)*100)-100), 2)
            elif Phenol > Phenol_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Phenol_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Phenol_Z2_A_el_f)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z2_A_el_f)*100)-100), 2)
                depV = True

            if Leitfähigkeit <= Leitfähigkeit_Z0_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Leitfähigkeit_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Leitfähigkeit_Z0_A_el_f)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z0_A_el_f)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z0_A_el_f and Leitfähigkeit <= Leitfähigkeit_Z1_1_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Leitfähigkeit_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Leitfähigkeit_Z1_1_A_el_f)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z0_A_el_f)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z1_1_A_el_f and Leitfähigkeit <= Leitfähigkeit_Z1_2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Leitfähigkeit_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Leitfähigkeit_Z1_2_A_el_f)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z1_1_A_el_f)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z1_2_A_el_f and Leitfähigkeit <= Leitfähigkeit_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Leitfähigkeit_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Leitfähigkeit_Z2_A_el_f)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z1_2_A_el_f)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z2_A_el_f and Recyclingbaustoffcheck_erweitert == False:
                z_Leitfähigkeit_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Leitfähigkeit_Z2_A_el_f)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z2_A_el_f)*100)-100), 2)
                depV = True
        # if Recyclingbaustoffcheck_erweitert == True:
            if As_el <= As_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_as_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(As_Z0_A_el)
                As_el_erhöhung = round((((As_el/As_Z0_A_el)*100)-100), 2)
            elif As_el > As_Z0_A_el and As_el <= As_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_as_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(As_Z1_1_A_el)
                As_el_erhöhung = round((((As_el/As_Z0_A_el)*100)-100), 2)
            elif As_el > As_Z1_1_A_el and As_el <= As_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_as_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(As_Z1_2_A_el)
                As_el_erhöhung = round((((As_el/As_Z1_1_A_el)*100)-100), 2)
            elif As_el > As_Z1_2_A_el and As_el <= As_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_as_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(As_Z2_A_el)
                As_el_erhöhung = round((((As_el/As_Z1_2_A_el)*100)-100), 2)
            elif As_el > As_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_as_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(As_Z2_A_el)
                As_el_erhöhung = round((((As_el/As_Z2_A_el)*100)-100), 2)
                depV = True

            if Pb_el <= Pb_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Pb_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Pb_Z0_A_el)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z0_A_el)*100)-100), 2)
            elif Pb_el > Pb_Z0_A_el and Pb_el <= Pb_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Pb_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Pb_Z1_1_A_el)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z0_A_el)*100)-100), 2)
            elif Pb_el > Pb_Z1_1_A_el and Pb_el <= Pb_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Pb_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Pb_Z1_2_A_el)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z1_1_A_el)*100)-100), 2)
            elif Pb_el > Pb_Z1_2_A_el and Pb_el <= Pb_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Pb_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Pb_Z2_A_el)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z1_2_A_el)*100)-100), 2)
            elif Pb_el > Pb_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Pb_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Pb_Z2_A_el)
                Pb_el_erhöhung = round((((Pb_el/Pb_Z2_A_el)*100)-100), 2)
                depV = True

            if Cd_el <= Cd_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cd_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Cd_Z0_A_el)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z0_A_el)*100)-100), 2)
            elif Cd_el > Cd_Z0_A_el and Cd_el <= Cd_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cd_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cd_Z1_1_A_el)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z0_A_el)*100)-100), 2)
            elif Cd_el > Cd_Z1_1_A_el and Cd_el <= Cd_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cd_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cd_Z1_2_A_el)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z1_1_A_el)*100)-100), 2)
            elif Cd_el > Cd_Z1_2_A_el and Cd_el <= Cd_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cd_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cd_Z2_A_el)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z1_2_A_el)*100)-100), 2)
            elif Cd_el > Cd_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cd_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cd_Z2_A_el)
                Cd_el_erhöhung = round((((Cd_el/Cd_Z2_A_el)*100)-100), 2)
                depV = True

            if Cr_el <= Cr_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cr_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Cr_Z0_A_el)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z0_A_el)*100)-100), 2)
            elif Cr_el > Cr_Z0_A_el and Cr_el <= Cr_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cr_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cr_Z1_1_A_el)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z0_A_el)*100)-100), 2)
            elif Cr_el > Cr_Z1_1_A_el and Cr_el <= Cr_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cr_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cr_Z1_2_A_el)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z1_1_A_el)*100)-100), 2)
            elif Cr_el > Cr_Z1_2_A_el and Cr_el <= Cr_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cr_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cr_Z2_A_el)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z1_2_A_el)*100)-100), 2)
            elif Cr_el > Cr_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cr_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cr_Z2_A_el)
                Cr_el_erhöhung = round((((Cr_el/Cr_Z2_A_el)*100)-100), 2)
                depV = True

            if Cu_el <= Cu_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cu_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Cu_Z0_A_el)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z0_A_el)*100)-100), 2)
            elif Cu_el > Cu_Z0_A_el and Cu_el <= Cu_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cu_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cu_Z1_1_A_el)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z0_A_el)*100)-100), 2)
            elif Cu_el > Cu_Z1_1_A_el and Cu_el <= Cu_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cu_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cu_Z1_2_A_el)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z1_1_A_el)*100)-100), 2)
            elif Cu_el > Cu_Z1_2_A_el and Cu_el <= Cu_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cu_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cu_Z2_A_el)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z1_2_A_el)*100)-100), 2)
            elif Cu_el > Cu_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cu_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cu_Z2_A_el)
                Cu_el_erhöhung = round((((Cu_el/Cu_Z2_A_el)*100)-100), 2)
                depV = True

            if Ni_el <= Ni_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Ni_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Ni_Z0_A_el)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z0_A_el)*100)-100), 2)
            elif Ni_el > Ni_Z0_A_el and Ni_el <= Ni_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Ni_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Ni_Z1_1_A_el)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z0_A_el)*100)-100), 2)
            elif Ni_el > Ni_Z1_1_A_el and Ni_el <= Ni_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Ni_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Ni_Z1_2_A_el)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z1_1_A_el)*100)-100), 2)
            elif Ni_el > Ni_Z1_2_A_el and Ni_el <= Ni_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Ni_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Ni_Z2_A_el)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z1_2_A_el)*100)-100), 2)
            elif Ni_el > Ni_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Ni_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Ni_Z2_A_el)
                Ni_el_erhöhung = round((((Ni_el/Ni_Z2_A_el)*100)-100), 2)
                depV = True

            if Hg_el <= Hg_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Hg_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Hg_Z0_A_el)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z0_A_el)*100)-100), 2)
            elif Hg_el > Hg_Z0_A_el and Hg_el <= Hg_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Hg_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Hg_Z1_1_A_el)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z0_A_el)*100)-100), 2)
            elif Hg_el > Hg_Z1_1_A_el and Hg_el <= Hg_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Hg_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Hg_Z1_2_A_el)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z1_1_A_el)*100)-100), 2)
            elif Hg_el > Hg_Z1_2_A_el and Hg_el <= Hg_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Hg_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Hg_Z2_A_el)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z1_2_A_el)*100)-100), 2)
            elif Hg_el > Hg_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Hg_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Hg_Z2_A_el)
                Hg_el_erhöhung = round((((Hg_el/Hg_Z2_A_el)*100)-100), 2)
                depV = True

            if Zn_el <= Zn_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Zn_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Zn_Z0_A_el)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z0_A_el)*100)-100), 2)
            elif Zn_el > Zn_Z0_A_el and Zn_el <= Zn_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Zn_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Zn_Z1_1_A_el)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z0_A_el)*100)-100), 2)
            elif Zn_el > Zn_Z1_1_A_el and Zn_el <= Zn_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Zn_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Zn_Z1_2_A_el)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z1_1_A_el)*100)-100), 2)
            elif Zn_el > Zn_Z1_2_A_el and Zn_el <= Zn_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Zn_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Zn_Z2_A_el)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z1_2_A_el)*100)-100), 2)
            elif Zn_el > Zn_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Zn_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Zn_Z2_A_el)
                Zn_el_erhöhung = round((((Zn_el/Zn_Z2_A_el)*100)-100), 2)
                depV = True
            # print("*********************CL********")
            # print(Cl_el)
            if Cl_el <= Cl_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cl_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Cl_Z0_A_el)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z0_A_el)*100)-100), 2)
            elif Cl_el > Cl_Z0_A_el and Cl_el <= Cl_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cl_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Cl_Z1_1_A_el)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z0_A_el)*100)-100), 2)
            elif Cl_el > Cl_Z1_1_A_el and Cl_el <= Cl_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cl_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Cl_Z1_2_A_el)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z1_1_A_el)*100)-100), 2)
            elif Cl_el > Cl_Z1_2_A_el and Cl_el <= Cl_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cl_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Cl_Z2_A_el)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z1_2_A_el)*100)-100), 2)
            elif Cl_el > Cl_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Cl_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Cl_Z2_A_el)
                Cl_el_erhöhung = round((((Cl_el/Cl_Z2_A_el)*100)-100), 2)
                depV = True
            # print("********SULFAT*****")
            # print(So4_el)
            if So4_el <= So4_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_So4_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(So4_Z0_A_el)
                So4_el_erhöhung = round((((So4_el/So4_Z0_A_el)*100)-100), 2)
            elif So4_el > So4_Z0_A_el and So4_el <= So4_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_So4_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(So4_Z1_1_A_el)
                So4_el_erhöhung = round((((So4_el/So4_Z0_A_el)*100)-100), 2)
            elif So4_el > So4_Z1_1_A_el and So4_el <= So4_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_So4_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(So4_Z1_2_A_el)
                So4_el_erhöhung = round((((So4_el/So4_Z1_1_A_el)*100)-100), 2)
            elif So4_el > So4_Z1_2_A_el and So4_el <= So4_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_So4_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(So4_Z2_A_el)
                So4_el_erhöhung = round((((So4_el/So4_Z1_2_A_el)*100)-100), 2)
            elif So4_el > So4_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_So4_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(So4_Z2_A_el)
                So4_el_erhöhung = round((((So4_el/So4_Z2_A_el)*100)-100), 2)
                depV = True

            if Phenol < Phenol_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Phenol_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [Z1.1]')
                O_Grenze_Z.append(Phenol_Z0_A_el)
                Phenol_erhöhung = round((((Phenol/Phenol_Z0_A_el)*100)-100), 2)
            elif Phenol == Phenol_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Phenol_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Phenol_Z1_1_A_el)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z0_A_el)*100)-100), 2)  # 10
            elif Phenol > Phenol_Z1_1_A_el and Phenol <= Phenol_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Phenol_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Phenol_Z1_2_A_el)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z1_1_A_el)*100)-100), 2)
            elif Phenol > Phenol_Z1_2_A_el and Phenol <= Phenol_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Phenol_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Phenol_Z2_A_el)
                Phenol_erhöhung = round(
                    (((Phenol/Phenol_Z1_2_A_el)*100)-100), 2)
            elif Phenol > Phenol_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Phenol_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Phenol_Z2_A_el)
                Phenol_erhöhung = round((((Phenol/Phenol_Z2_A_el)*100)-100), 2)
                depV = True

            pH_depV_below = False
            pH_depV = False
            pH_depV_above = False
            #pH_carb = 13
            if pH_carb:
                pH_carb = float(pH_carb)
                if Recyclingbaustoffcheck_erweitert == True or False:
                    if pH_carb < 7.0:
                        z_pH_el = 4
                        z_ph_el_std = 3
                        pH_depV_below = True
                        pH_depV = True
                    elif pH_carb >= 7.0 and pH_carb <= 12.5:
                        z_ph_el_std = 0
                        pH_depV = False
                        print("testetsetestefvjhvvsejv")
                    elif pH_carb > 12.5:
                        z_pH_el = 4
                        z_ph_el_std = 3
                        pH_depV_above = True
                        pH_depV = True
            elif pH:
                pH = float(pH)
                if Recyclingbaustoffcheck_erweitert == True or False:
                    if pH < 7.0:
                        z_pH_el = 4
                        z_ph_el_std = 3
                        pH_depV_below = True
                        pH_depV = True
                    elif pH >= 7.0 and pH <= 12.5:
                        z_ph_el_std = 0
                        pH_depV = False
                    elif pH > 12.5:
                        z_pH_el = 4
                        z_ph_el_std = 3
                        pH_depV = True
                        pH_depV_above = True
            if Leitfähigkeit <= Leitfähigkeit_Z0_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Leitfähigkeit_A_el = 0
                z_el_A.append(0)
                Z_prior.append('[Z0]')
                Z_next.append(' -> [>Z1.1]')
                O_Grenze_Z.append(Leitfähigkeit_Z0_A_el)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z0_A_el)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z0_A_el and Leitfähigkeit <= Leitfähigkeit_Z1_1_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Leitfähigkeit_A_el = 1.1
                z_el_A.append(1.1)
                Z_prior.append('[Z1.1]')
                Z_next.append(' -> [Z1.2]')
                O_Grenze_Z.append(Leitfähigkeit_Z1_1_A_el)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z0_A_el)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z1_1_A_el and Leitfähigkeit <= Leitfähigkeit_Z1_2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Leitfähigkeit_A_el = 1.2
                z_el_A.append(1.2)
                Z_prior.append('[Z1.2]')
                Z_next.append(' -> [Z2]')
                O_Grenze_Z.append(Leitfähigkeit_Z1_2_A_el)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z1_1_A_el)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z1_2_A_el and Leitfähigkeit <= Leitfähigkeit_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Leitfähigkeit_A_el = 2
                z_el_A.append(2)
                Z_prior.append('[Z2]')
                Z_next.append(' -> [>Z2]')
                O_Grenze_Z.append(Leitfähigkeit_Z2_A_el)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z1_2_A_el)*100)-100), 2)
            elif Leitfähigkeit > Leitfähigkeit_Z2_A_el and Recyclingbaustoffcheck_erweitert == True:
                z_Leitfähigkeit_A_el = 3
                z_el_A.append(9)
                Z_prior.append('[>Z2]')
                Z_next.append(' -> [DK]')
                O_Grenze_Z.append(Leitfähigkeit_Z2_A_el)
                Leitfähigkeit_erhöhung = round(
                    (((Leitfähigkeit/Leitfähigkeit_Z2_A_el)*100)-100), 2)
                depV = True

            Z_Rec_false_SM = []
            if Recyclingbaustoffcheck_erweitert == True:
                if Arsen_A_erhöhung < 0 or Arsen_A_erhöhung == 0:
                    Arsen_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Arsen_A_erhöhung_plot = str(
                        Arsen_A_erhöhung)+" % "  # +Z_prior[0]
                    Z_Rec_false_SM.append(Z_prior[0])
                if Blei_A_erhöhung < 0 or Blei_A_erhöhung == 0:
                    Blei_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Blei_A_erhöhung_plot = str(
                        Blei_A_erhöhung)+" % "  # +Z_prior[1]
                    Z_Rec_false_SM.append(Z_prior[1])
                if Cadmium_A_erhöhung < 0 or Cadmium_A_erhöhung == 0:
                    Cadmium_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Cadmium_A_erhöhung_plot = str(
                        Cadmium_A_erhöhung)+" % "  # +Z_prior[2]
                    Z_Rec_false_SM.append(Z_prior[2])
                if Chrom_A_erhöhung < 0 or Chrom_A_erhöhung == 0:
                    Chrom_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Chrom_A_erhöhung_plot = str(
                        Chrom_A_erhöhung)+" % "  # +Z_prior[3]
                    Z_Rec_false_SM.append(Z_prior[3])
                if Kupfer_A_erhöhung < 0 or Kupfer_A_erhöhung == 0:
                    Kupfer_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Kupfer_A_erhöhung_plot = str(
                        Kupfer_A_erhöhung)+" % "  # +Z_prior[4]
                    Z_Rec_false_SM.append(Z_prior[4])
                if Nickel_A_erhöhung < 0 or Nickel_A_erhöhung == 0:
                    Nickel_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Nickel_A_erhöhung_plot = str(
                        Nickel_A_erhöhung)+" % "  # +Z_prior[5]
                    Z_Rec_false_SM.append(Z_prior[5])

                if Quecksilber_A_erhöhung < 0 or Quecksilber_A_erhöhung == 0:
                    Quecksilber_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Quecksilber_A_erhöhung_plot = str(
                        Quecksilber_A_erhöhung)+" % "  # +Z_prior[6]
                    Z_Rec_false_SM.append(Z_prior[6])

                if Zink_A_erhöhung < 0 or Zink_A_erhöhung == 0:
                    Zink_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Zink_A_erhöhung_plot = str(
                        Zink_A_erhöhung)+" % "  # +Z_prior[7]
                    Z_Rec_false_SM.append(Z_prior[7])
                if Eox_A_erhöhung < 0 or Eox_A_erhöhung == 0:
                    Eox_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Eox_A_erhöhung_plot = str(
                        Eox_A_erhöhung)+" % "  # +Z_prior[8]
                    Z_Rec_false_SM.append(Z_prior[8])

                if Kw_A_erhöhung < 0 or Kw_A_erhöhung == 0:
                    Kw_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Kw_A_erhöhung_plot = str(
                        Kw_A_erhöhung)+" % "+str(Kw_A)+" "  # + Z_prior[9]
                    Z_Rec_false_SM.append(Z_prior[9])

                if Pak16_A_erhöhung < 0 or Pak16_A_erhöhung == 0:
                    Pak16_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Pak16_A_erhöhung_plot = str(
                        Pak16_A_erhöhung)+" % "  # +Z_prior[10]
                    Z_Rec_false_SM.append(Z_prior[10])

                if Summe_PCB_A_erhöhung < 0 or Summe_PCB_A_erhöhung == 0:
                    Summe_PCB_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Summe_PCB_A_erhöhung_plot = str(
                        Summe_PCB_A_erhöhung)+" % "  # +Z_prior[11]
                    Z_Rec_false_SM.append(Z_prior[11])

            elif Recyclingbaustoffcheck_erweitert == False:
                if Arsen_A_erhöhung < 0 or Arsen_A_erhöhung == 0:
                    Arsen_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Arsen_A_erhöhung_plot = str(
                        Arsen_A_erhöhung)+" % "  # [>Z0]
                    Z_Rec_false_SM.append('>Z0')
                if Blei_A_erhöhung < 0 or Blei_A_erhöhung == 0:
                    Blei_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Blei_A_erhöhung_plot = str(
                        Blei_A_erhöhung)+" % "
                    Z_Rec_false_SM.append('>Z0')
                if Cadmium_A_erhöhung < 0 or Cadmium_A_erhöhung == 0:
                    Cadmium_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Cadmium_A_erhöhung_plot = str(
                        Cadmium_A_erhöhung)+" % "
                    Z_Rec_false_SM.append('>Z0')
                if Chrom_A_erhöhung < 0 or Chrom_A_erhöhung == 0:
                    Chrom_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Chrom_A_erhöhung_plot = str(
                        Chrom_A_erhöhung)+" % "
                    Z_Rec_false_SM.append('>Z0')
                if Kupfer_A_erhöhung < 0 or Kupfer_A_erhöhung == 0:
                    Kupfer_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Kupfer_A_erhöhung_plot = str(
                        Kupfer_A_erhöhung)+" % "
                    Z_Rec_false_SM.append('>Z0')
                if Nickel_A_erhöhung < 0 or Nickel_A_erhöhung == 0:
                    Nickel_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Nickel_A_erhöhung_plot = str(
                        Nickel_A_erhöhung)+" % "
                    Z_Rec_false_SM.append('>Z0')
                if Quecksilber_A_erhöhung < 0 or Quecksilber_A_erhöhung == 0:
                    Quecksilber_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Quecksilber_A_erhöhung_plot = str(
                        Quecksilber_A_erhöhung)+" % "
                    Z_Rec_false_SM.append('>Z0')
                if Zink_A_erhöhung < 0 or Zink_A_erhöhung == 0:
                    Zink_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                elif Zink_A_erhöhung > 0:
                    Zink_A_erhöhung_plot = str(Zink_A_erhöhung)+" % "
                    Z_Rec_false_SM.append('>Z0')

                else:
                    Zink_A_erhöhung_plot = str(
                        Zink_A_erhöhung)+" % "  # +Z_prior[7]
                    Z_Rec_false_SM.append(Z_prior[7])
                if Eox_A_erhöhung < 0 or Eox_A_erhöhung == 0:
                    Eox_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Eox_A_erhöhung_plot = str(
                        Eox_A_erhöhung)+" % "  # +Z_prior[8]
                    Z_Rec_false_SM.append(Z_prior[8])
                if Kw_A_erhöhung < 0 or Kw_A_erhöhung == 0:
                    Kw_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Kw_A_erhöhung_plot = str(
                        Kw_A_erhöhung)+" % "  # +Z_prior[9]
                    Z_Rec_false_SM.append(Z_prior[9])

                if Pak16_A_erhöhung < 0 or Pak16_A_erhöhung == 0:
                    Pak16_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Pak16_A_erhöhung_plot = str(
                        Pak16_A_erhöhung)+" % "  # +Z_prior[10]
                    Z_Rec_false_SM.append(Z_prior[10])

                if Summe_PCB_A_erhöhung < 0 or Summe_PCB_A_erhöhung == 0:
                    Summe_PCB_A_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Summe_PCB_A_erhöhung_plot = str(
                        Summe_PCB_A_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[11])
            # ELUATE
            if Recyclingbaustoffcheck_erweitert == True or Recyclingbaustoffcheck_erweitert == False:
                if As_el_erhöhung < 0 or As_el_erhöhung == 0:
                    As_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    As_el_erhöhung_plot = str(
                        As_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[12])

                if Pb_el_erhöhung < 0 or Pb_el_erhöhung == 0:
                    Pb_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Pb_el_erhöhung_plot = str(
                        Pb_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[13])

                if Cd_el_erhöhung < 0 or Cd_el_erhöhung == 0:
                    Cd_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Cd_el_erhöhung_plot = str(
                        Cd_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[14])

                if Cr_el_erhöhung < 0 or Cr_el_erhöhung == 0:
                    Cr_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Cr_el_erhöhung_plot = str(
                        Cr_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[15])

                if Cu_el_erhöhung < 0 or Cu_el_erhöhung == 0:
                    Cu_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Cu_el_erhöhung_plot = str(
                        Cu_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[16])

                if Ni_el_erhöhung < 0 or Ni_el_erhöhung == 0:
                    Ni_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Ni_el_erhöhung_plot = str(
                        Ni_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[17])

                if Hg_el_erhöhung < 0 or Hg_el_erhöhung == 0:
                    Hg_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Hg_el_erhöhung_plot = str(
                        Hg_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[18])

                if Zn_el_erhöhung < 0 or Zn_el_erhöhung == 0:
                    Zn_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Zn_el_erhöhung_plot = str(
                        Zn_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[19])

                if Cl_el_erhöhung < 0 or Cl_el_erhöhung == 0:
                    Cl_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Cl_el_erhöhung_plot = str(
                        Cl_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[20])

                if So4_el_erhöhung < 0 or So4_el_erhöhung == 0:
                    So4_el_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    So4_el_erhöhung_plot = str(
                        So4_el_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[21])

                if Phenol_erhöhung < 0 or Phenol_erhöhung == 0:
                    Phenol_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Phenol_erhöhung_plot = str(
                        Phenol_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[22])

                if Leitfähigkeit_erhöhung < 0 or Leitfähigkeit_erhöhung == 0:
                    Leitfähigkeit_erhöhung_plot = ""
                    Z_Rec_false_SM.append('')
                else:
                    Leitfähigkeit_erhöhung_plot = str(
                        Leitfähigkeit_erhöhung)+" % "
                    Z_Rec_false_SM.append(Z_prior[23])

            if Recyclingbaustoffcheck_erweitert == False:
                if z_sm_A[8] or z_sm_A[9] or z_sm_A[10] or z_sm_A[11] == 1.1 or z_sm_A[8] or z_sm_A[9] or z_sm_A[10] or z_sm_A[11] == 1.2 or z_sm_A[8] or z_sm_A[9] or z_sm_A[10] or z_sm_A[11] == 2 or z_sm_A[8] or z_sm_A[9] or z_sm_A[10] or z_sm_A[11] == 3 or z_sm_A[8] or z_sm_A[9] or z_sm_A[10] or z_sm_A[11] == 9:
                    max_z_sm_A = max(
                        list([z_sm_A[8], z_sm_A[9], z_sm_A[10], z_sm_A[11]]))

                else:
                    max_z_sm_A = max(list([z_sm_A[0], z_sm_A[1], z_sm_A[2], z_sm_A[3],
                                           z_sm_A[4], z_sm_A[5], z_sm_A[6], z_sm_A[7], z_sm_A[8], z_sm_A[9], z_sm_A[10], z_sm_A[11]]))
            elif Recyclingbaustoffcheck_erweitert == True:
                max_z_sm_A = max(list([z_sm_A[0], z_sm_A[1], z_sm_A[2], z_sm_A[3],
                                       z_sm_A[4], z_sm_A[5], z_sm_A[6], z_sm_A[7], z_sm_A[8], z_sm_A[9], z_sm_A[10], z_sm_A[11]]))

            if Recyclingbaustoffcheck_erweitert == True or Recyclingbaustoffcheck_erweitert == False:
                max_z_el_A = max(list([z_el_A[0], z_el_A[1], z_el_A[2],
                                       z_el_A[3], z_el_A[4], z_el_A[5], z_el_A[6], z_el_A[7], z_el_A[8], z_el_A[9], z_el_A[10], z_el_A[11]]))

            if max_z_el_A == 0:
                zuordnungswert_el_A = 'Z0'
                farbe_grenzlinie = "#088A08"
                sucess_zuordnung_el_A = 1
            elif max_z_el_A == 1.1:
                zuordnungswert_el_A = 'Z1.1'
                farbe_grenzlinie = "#088A4B"
                sucess_zuordnung_el_A = 2
            elif max_z_el_A == 1.2:
                zuordnungswert_el_A = 'Z1.2'
                farbe_grenzlinie = "#FF4000"
                sucess_zuordnung_el_A = 3
            elif max_z_el_A == 2:
                zuordnungswert_el_A = 'Z2'
                farbe_grenzlinie = "#FF0000"
                sucess_zuordnung_el_A = 4
            elif max_z_el_A == 9:
                zuordnungswert_el_A = 'Z2'
                farbe_grenzlinie = "#FF0000"
                sucess_zuordnung_el_A = 4

            if max_z_sm_A == 9 and Recyclingbaustoffcheck_erweitert == True:
                zuordnungswert_sm_A = '>Z2'
                farbe_grenzlinie = "#FF0000"
                sucess_zuordnung_A = 0
            elif max_z_sm_A == 3 and Recyclingbaustoffcheck_erweitert == False:
                zuordnungswert_sm_A = '>Z2'
                farbe_grenzlinie = "#FF0000"
                sucess_zuordnung_A = 0
            elif max_z_sm_A == 9 and Recyclingbaustoffcheck_erweitert == False:
                zuordnungswert_sm_A = '>Z0'
                farbe_grenzlinie = "#088A4B"
                sucess_zuordnung_A = 5
            elif max_z_sm_A == 1.1 and Recyclingbaustoffcheck_erweitert == False:
                zuordnungswert_sm_A = 'Z1.1'
                farbe_grenzlinie = "#088A4B"
                sucess_zuordnung_A = 2
            elif max_z_sm_A == 1.2 and Recyclingbaustoffcheck_erweitert == False:
                zuordnungswert_sm_A = 'Z1.2'
                farbe_grenzlinie = "#FF4000"
                sucess_zuordnung_A = 3
            elif max_z_sm_A == 2 and Recyclingbaustoffcheck_erweitert == False:
                zuordnungswert_sm_A = 'Z2'
                farbe_grenzlinie = "#FF0000"
                sucess_zuordnung_A = 4

            elif max_z_sm_A == 0:
                zuordnungswert_sm_A = 'Z0'
                farbe_grenzlinie = "#088A08"
                sucess_zuordnung_A = 1
            elif max_z_sm_A == 1.1:
                zuordnungswert_sm_A = 'Z1.1'
                farbe_grenzlinie = "#088A4B"
                sucess_zuordnung_A = 2
            elif max_z_sm_A == 1.2:
                zuordnungswert_sm_A = 'Z1.2'
                farbe_grenzlinie = "#FF4000"
                sucess_zuordnung_A = 3
            elif max_z_sm_A == 2:
                zuordnungswert_sm_A = 'Z2'
                farbe_grenzlinie = "#FF0000"
                sucess_zuordnung_A = 4

            Z_Klassifikation_feststoff = z_sm_A

            bestimmende_Parameter_liste_feststoff = ['Arsen ', 'Blei ', 'Cadmium ', 'Chrom ', 'Kupfer ',
                                                     'Nickel ', 'Quecksilber ', 'Zink ', 'EOX ', 'Kohlenwasserstoffe ', 'PAK16 ', 'PCB ']

            bestimmende_Parameter_feststoff = []
            for i in range(0, len(bestimmende_Parameter_liste_feststoff)):
                if max_z_sm_A == Z_Klassifikation_feststoff[i]:

                    bestimmender_parameter_feststoff = bestimmende_Parameter_liste_feststoff[i]
                    bestimmende_Parameter_feststoff.append(
                        bestimmender_parameter_feststoff)

            for i in range(0, len(bestimmende_Parameter_feststoff)):
                if max_z_sm_A == 0:
                    bestimmende_Parameter_feststoff = "Z0"
                else:
                    bestimmende_Parameter_feststoff = "".join(
                        bestimmende_Parameter_feststoff)

            Z_Klassifikation_feststoff = "Z" + \
                str(max(Z_Klassifikation_feststoff))

            if Z_Klassifikation_feststoff == "Z0":
                Z_Klassifikation_feststoff = "Z0"
                bem_feststoff = "Bemerkungen: "
            elif Z_Klassifikation_feststoff == "Z1.1":
                Z_Klassifikation_feststoff = "Z1.1"
                bem_feststoff = "Bemerkungen: " + \
                    str(bestimmende_Parameter_feststoff)
            elif Z_Klassifikation_feststoff == "Z1.2":
                Z_Klassifikation_feststoff = "Z1.2"
                bem_feststoff = "Bemerkungen: " + \
                    str(bestimmende_Parameter_feststoff)
            elif Z_Klassifikation_feststoff == "Z2":
                Z_Klassifikation_feststoff = "Z2"
                bem_feststoff = "Bemerkungen: " + \
                    str(bestimmende_Parameter_feststoff)

            elif Z_Klassifikation_feststoff == "Z9" or "Z9.0":
                Z_Klassifikation_feststoff = ">Z2"
                bem_feststoff = "Bemerkungen: " + \
                    str(bestimmende_Parameter_feststoff)
            else:
                bem_feststoff = "Bemerkungen: " + \
                    str(bestimmende_Parameter_feststoff)

            Z_Klassifikation = z_sm_A+z_el_A
            Z_Klassifikation_rec_False = Z_Klassifikation
            # print("max Z Klas.")
            # print(max(Z_Klassifikation))
            # print(Z_Klassifikation)
            bestimmende_Parameter_liste = ['Arsen ', 'Blei ', 'Cadmium ', 'Chrom ', 'Kupfer ', 'Nickel ', 'Quecksilber ', 'Zink ', 'EOX ', 'Kohlenwasserstoffe ', 'PAK16 ', 'PCB ',
                                           'Arsen (Eluat) ', 'Blei (Eluat) ', 'Cadmium (Eluat) ', 'Chrom (Eluat) ', 'Kupfer (Eluat) ', 'Nickel (Eluat) ', 'Quecksilber (Eluat) ', 'Zink (Eluat) ', 'Chlorid ', 'Sulfat ', 'Phenol ', 'Leitfähigkeit ']
            bestimmende_Parameter = []
            # print("maxcheck")
            # print(max(Z_Klassifikation))
            if Recyclingbaustoffcheck_erweitert == True:
                if max(Z_Klassifikation) == 0:
                    # print("JUHCUHCUHu")
                    bestimmende_Parameter = ""
                    Z_Klassifikation = "Z0"
                    bem = "Bemerkungen: " + \
                        str(bestimmende_Parameter)
                else:
                    for i in range(0, len(bestimmende_Parameter_liste)):  # 0, len(Z_Klassifikation)) and range(
                        if max(Z_Klassifikation) == Z_Klassifikation[i]:
                            bestimmender_parameter = bestimmende_Parameter_liste[i]
                            bestimmende_Parameter.append(
                                bestimmender_parameter)
                    for i in range(0, len(bestimmende_Parameter)):
                        if max(Z_Klassifikation) == 0:
                            bestimmende_Parameter = ""
                        else:
                            bestimmende_Parameter = "".join(
                                bestimmende_Parameter)
                    if max(Z_Klassifikation) == 1:
                        Z_Klassifikation = "Z1"
                    elif max(Z_Klassifikation) == 1.1:
                        Z_Klassifikation = "Z1.1"
                    elif max(Z_Klassifikation) == 1.2:
                        #print("test")
                        Z_Klassifikation = "Z1.2"
                    elif max(Z_Klassifikation) == 2:
                        Z_Klassifikation = "Z2"
                    elif max(Z_Klassifikation) == 9:
                        Z_Klassifikation = ">Z2"
                    bem = "Bemerkungen: " + \
                        str(bestimmende_Parameter)

            

            # print("Test Z Klass")
            # print(Z_Klassifikation)

            # if max(Z_Klassifikation) == 2.0 or 0.0:
            #     Z_Klassifikation = "Z"+str((max(Z_Klassifikation)))
            # else:
            #     Z_Klassifikation = "Z"+str(float(max(Z_Klassifikation)))

            false_sm_z_klassifikation = []
            false_sm_z_klassifikation.append(z_eox_A)
            false_sm_z_klassifikation.append(z_kw_A)
            false_sm_z_klassifikation.append(z_pak16_A)

            bestimmende_Parameter_liste_false_rec = ['Arsen ', 'Blei ', 'Cadmium ', 'Chrom ', 'Kupfer ', 'Nickel ', 'Quecksilber ', 'Zink ', 'EOX ', 'Kohlenwasserstoffe ', 'PAK16 ', 'PCB ',
                                                     'Arsen (Eluat) ', 'Blei (Eluat) ', 'Cadmium (Eluat) ', 'Chrom (Eluat) ', 'Kupfer (Eluat) ', 'Nickel (Eluat) ', 'Quecksilber (Eluat) ', 'Zink (Eluat) ', 'Chlorid ', 'Sulfat ', 'Phenol ', 'Leitfähigkeit ']

            bestimmende_Parameter_rec_false = []
            if Recyclingbaustoffcheck_erweitert == False:
                if max(Z_Klassifikation_rec_False) == 0:
                    bestimmende_Parameter_rec_false = ""
                else:
                    for i in range(0, len(bestimmende_Parameter_liste_false_rec)):  # 0, len(Z_Klassifikation)) and range(
                        if max(Z_Klassifikation_rec_False) == Z_Klassifikation_rec_False[i]:
                            bestimmender_parameter = bestimmende_Parameter_liste_false_rec[i]
                            bestimmende_Parameter_rec_false.append(
                                bestimmender_parameter)
                            Z_Klassifikation_rec_False[i] = 0
                            if max(Z_Klassifikation_rec_False) == Z_Klassifikation_rec_False[i]:
                                bestimmender_parameter = bestimmende_Parameter_liste_false_rec[i]
                                bestimmende_Parameter_rec_false.append(
                                    bestimmender_parameter)
                    for i in range(0, len(bestimmende_Parameter_rec_false)):
                        if max(Z_Klassifikation_rec_False) == 0:
                            bestimmende_Parameter_rec_false = ""
                        else:
                            bestimmende_Parameter_rec_false = "".join(
                                bestimmende_Parameter_rec_false)
                    bestimmende_Parameter = bestimmende_Parameter_rec_false

            false_rest = []
            false_rest = [z_Cl_A_el, z_Leitfähigkeit_A_el,
                          z_Phenol_A_el, z_So4_A_el]  # H_depV == False
            if Recyclingbaustoffcheck_erweitert == False:
                if Z_Klassifikation == "Z9" or "Z9.0" and Recyclingbaustoffcheck_erweitert == False:
                    if (z_as_A or z_pb_A or z_cr_A or z_cd_A or z_cu_A or z_hg_A or z_zn_A or z_ni_A) > max(z_el_A):
                        if (z_as_A or z_pb_A or z_cr_A or z_cd_A or z_cu_A or z_hg_A or z_zn_A or z_ni_A) == 9 and pH_depV == False:
                            if max(false_rest) == 0:
                                Z_Klassifikation = ">Z0"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_rest) == 1.1:
                                Z_Klassifikation = "Z1.1"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_rest) == 1.2:
                                Z_Klassifikation = "Z1.2"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_rest) == 2:
                                Z_Klassifikation = "Z2"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_rest) == 3:
                                Z_Klassifikation = ">Z2"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                        elif (z_as_A or z_pb_A or z_cr_A or z_cd_A or z_cu_A or z_hg_A or z_zn_A or z_ni_A) == 9 and pH_depV == True:
                            Z_Klassifikation = ">Z2"
                            bem = "Bemerkungen: " + str(bestimmende_Parameter)
                        elif max(false_sm_z_klassifikation) == 0:
                            Z_Klassifikation = "Z0"
                            bem = "Bemerkungen: " + \
                                str(bestimmende_Parameter)
                        elif max(false_sm_z_klassifikation) == max(z_el_A):
                            if max(false_sm_z_klassifikation) == 1.1:
                                Z_Klassifikation = "Z1.1"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_sm_z_klassifikation) == 1.2:
                                Z_Klassifikation = "Z1.2"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_sm_z_klassifikation) == 2 or 2.0:
                                Z_Klassifikation = "Z2"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_sm_z_klassifikation) == 3 or 3.0:
                                Z_Klassifikation = ">Z2"
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                        elif max(false_sm_z_klassifikation) < max(z_el_A):
                            if max(z_el_A) == 1.1:
                                Z_Klassifikation = "Z1.1"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(z_el_A) == 1.2:
                                Z_Klassifikation = "Z1.2"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(z_el_A) == 2 or 2.0:
                                Z_Klassifikation = "Z2"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(z_el_A) == 3 or 3.0:
                                Z_Klassifikation = ">Z2"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                        elif max(false_sm_z_klassifikation) > max(z_el_A):
                            if max(false_sm_z_klassifikation) == 1.1:
                                Z_Klassifikation = "Z1.1"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_sm_z_klassifikation) == 1.2:
                                Z_Klassifikation = "Z1.2"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_sm_z_klassifikation) == 2 or 2.0:
                                Z_Klassifikation = "Z2"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                            elif max(false_sm_z_klassifikation) == 3 or 3.0:
                                Z_Klassifikation = ">Z2"
                                print(Z_Klassifikation)
                                bem = "Bemerkungen: " + \
                                    str(bestimmende_Parameter)
                        else:
                            Z_Klassifikation = ">Z0"
                            bem = "Bemerkungen: " + \
                                str(bestimmende_Parameter)
                    elif z_as_A or z_pb_A or z_cr_A or z_cd_A or z_cu_A or z_hg_A or z_zn_A or z_ni_A < z_Leitfähigkeit_A_el or z_So4_A_el or z_as_A_el or z_Pb_A_el or z_Cr_A_el or z_Cd_A_el or z_Cu_A_el or z_Ni_A_el or z_Zn_A_el or z_ph_el_std:
                        if max_z_el_A == 1.1:
                            Z_Klassifikation = "Z1.1"
                            bem = "Bemerkungen: " + \
                                str(bestimmende_Parameter)
                        elif max_z_el_A == 1.2:
                            Z_Klassifikation = "Z1.2"
                            bem = "Bemerkungen: " + \
                                str(bestimmende_Parameter)
                        elif max_z_el_A == 2:
                            Z_Klassifikation = "Z2"
                            bem = "Bemerkungen: " + \
                                str(bestimmende_Parameter)
                        elif max_z_el_A == 9:
                            Z_Klassifikation = ">Z2"
                            bem = "Bemerkungen: " + \
                                str(bestimmende_Parameter)
                else:
                    bem = "Bemerkungen: " + \
                        str(bestimmende_Parameter)
            if Recyclingbaustoffcheck_erweitert== True or Recyclingbaustoffcheck_erweitert==False:
                print(type(bestimmende_Parameter))
                if pH_depV==True and Z_Klassifikation ==">Z2":
                    bestimmende_Parameter=bestimmender_parameter+" pH"


            st.markdown('**Probenbezeichnung**: ' + Probenbezeichnung_A)

            if Z_Klassifikation == "Z0":
                #print("eieieiei")
                st.success("Die Probe " + str(Probenbezeichnung_A)+" wird gemäß dem Umfang der Analytik der Einbauklasse " + str(Z_Klassifikation) +
                           " (Uneingeschränkter Einbau) zugeordnet. \n\nDer Einbau gilt somit ausschließlich für die Verwertung in bodenähnlichen Anwendungen (Verfüllung von Abgrabungen und Abfallverwertung im Landschaftsbau außerhalb von Bauwerken).")
                st.markdown('**Bestimmende Parameter: **' +
                            bestimmende_Parameter)

            elif Z_Klassifikation == ">Z0":
                st.error("Die Probe " + str(Probenbezeichnung_A)+" wird gemäß dem Umfang der Analytik der Einbauklasse " + str(Z_Klassifikation) +
                         " zugeordnet (vgl. "+bauschuttquelle2003+", S. 54 ff.). Für die Schwermetalle Arsen, Blei, Cadmium, Chrom (gesamt), Kupfer, Nickel, Quecksilber und Zink werden für Massengehalte oberhalb der Z0-Grenze gem. LAGA Bauschutt keine Zuordnungswerte differenziert. Die Zuordnung resultiert weiterhin aus der o.g. Option das Material nicht für Rekultivierungszwecke oder als Einbaumaterial zu verwenden.")

            elif Z_Klassifikation == "Z1.1":
                st.warning("Die Probe " + str(Probenbezeichnung_A)+" wird gemäß dem Umfang der Analytik der Einbauklasse " + str(Z_Klassifikation) +
                           " (Eingeschränkter offener Einbau) zugeordnet. \n\nBei Einhaltung dieser Werte ist selbst unter ungünstigen hydrogeologischen Voraussetzungen davon auszugehen, dass keine nachteiligen Veränderungen des Grundwassers auftreten.")
                st.markdown('**Bestimmende Parameter: **' +
                            bestimmende_Parameter)

            elif Z_Klassifikation == "Z1.2":
                st.warning("Die Probe " + str(Probenbezeichnung_A)+" wird gemäß dem Umfang der Analytik der Einbauklasse " + str(Z_Klassifikation) + " (Eingeschränkter offener Einbau) zugeordnet. \n\nEin Einbau in hydrogeologisch ungünstigen Gebieten ist somit ausgeschlossen. Dennoch kann ein Einbau auch hier erfolgen, sofern dies landesspezifisch festgelegt ist und die Böden nachweislich eine Vorbelastung des Bodens > Z1.1 aufweisen (Verschlechterungsverbot). Sollten die hydrogeologisch günstigen Gebiete durch die zuständige Behörde nicht verbindlich festgelegt wurden sein, müssen die genehmigenden Behörden die geforderten günstigen Standorteigenschaften durch ein Gutachten nachweisen. Ein hohes Rückhaltevermögen ist i.d.R. bei mindestens 2 m mächtigen bindigen Deckschichten gegeben (vgl. "+bauschuttquelle2003+', S. 54 f.)')
                st.markdown('**Bestimmende Parameter: **' +
                            bestimmende_Parameter)

            elif Z_Klassifikation == "Z2" or "Z2.0":
                st.error("Die Probe " + str(Probenbezeichnung_A)+" wird gemäß dem Umfang der Analytik der Einbauklasse " + str(Z_Klassifikation) +
                         " (Eingeschränkter Einbau mit definierten technischen Sicherungsmaßnahmen) zugeordnet (vgl. "+bauschuttquelle2003+', S. 54 ff.)')
                st.markdown('**Bestimmende Parameter: **' +
                            bestimmende_Parameter)

            st.markdown('**Entnahmedatum**: ' + Entnahmedatum_A)
            st.markdown('**Prüfberichtnummer**: '+Prüfberichtnummer_A +
                        " ["+str(len(Prüfberichtnummer_A))+"-stellig]")
            st.markdown('**Datum des Prüfberichtes**: '+Datum_A)

        with st.beta_expander("Wichtige Hinweise"):
            st.markdown(
                '**Wichtige Hinweise**')

            hinweise = 0
            if Z_Klassifikation==">Z2":
                hinweise=1
                st.markdown("**Kritische Überschreitung**: Die Zuordnungswerte eines oder mehrerer Analyseparameter überschreiten den oberen Grenzbereich der Z2-Klassifikation uind liegen somit im Anwendungsbereich der Deponieverordnung (2009).")
            if Z_Klassifikation=="Z2" or Z_Klassifikation==">Z2":
                hinweise=1
                st.markdown("**Hohe Zuordnungswerte**: Die Zuordnungswerte einer oder mehrerer Analyseparameter sind mit einer Zuordnung zur Klassifikation von mindestens Z2 deutlich erhöht.")
            if Kw_A > 300:
                hinweise = 1
                st.markdown('**Kohlenwasserstoffe**: Kohlenwasserstoffe liegen oberhalb des Wertes von 300 mg/kg. Gemäß LAGA 1997 Bauschutt gilt hierfür, dass Überschreitungen, die auf Asphaltanteile zurückzuführen sind, kein Ausschlusskriterium für eine Zuordnung zu den entsprechenden Klassifikationen darstellen.')
            if Pak16_A > 15:
                hinweise = 1
                st.markdown(
                    '**PAK n. EPA / PAK16**: Der Summenwert der 16 PAK n. EPA liegt oberhalb von 15 mg/kg (Grenzwert Z1.2 -> Z2). Der Bereich der Zuordnung zu Z2-Material (Grenzwert [Z2]-> [>Z2] von 75 mg/kg auf 100 mg/kg) lässt sich unter folgenden Voraussetzungen erweitern: [1] Die erhöhten PAK-Gehalte sind auf pechhaltige Anteile zurückzuführen, [2] Es handelt sich um Baumaßnahmen im klassifizierten Straßenoberbau bzw. Verkehrsflächenoberbau (ausgenommen Wirtschaftswege), [3] Es handelt sich um eine größere Baumaßnahme (Volumen des eingebauten Recyclingbaustoffes > 500 m3), [4] Es handelt sich um Flächen, auf denen nicht mit häufigen Aufbrüchen gerechnet werden muss, [5] Die Recyclinganlage unterliegt einer regelmäßigen Güteüberwachung.')
            if hinweise == 0:
                st.markdown(
                    "Es bestehen gem. der Massenkonzentrationen an PAK oder KW keine besonderen Hinweise.")
        # with st.beta_expander:("Wichtige Hinweise:")
        # st.write("Im Folgenden sind potentiell wichtige Hinweise bzgl. der ermittelten Zuordnungswerte der jeweiligen Einzelparameter aufgeführt:")
        # if Kw_A>300:
        #             st.markdown('Kohlenwasserstoffe: Kohlenwasserstoffe liegen oberhalb des Wertes von 300 mg/kg. Gemäß LAGA 1997 Bauschutt gilt hierfür, dass Überschreitungen, die auf Asphaltanteile zurückzuführen sind, kein Ausschlusskriterium für eine Zuordnung zu den entsprechenden Klassifikationen darstellen.')
        # if Pak16_A>5:
        #             st.markdown('PAK n. EPA / PAK16: Der Summenwert der 16 PAK n. EPA liegt oberhalb von 5 mg/kg. In Einzelfällen kann für die Zuordnung zu Z1.1, Z1.2, Z2 sowie >Z2 abgewichen werden.')
        # if udb_conv_as==True:
        #     Arsen_A=''
        # if udb_conv_pb:
        #     Blei_A=''
        # if udb_conv_cd:
        #     Cadmium_A=''
        # if udb_conv_cr:
        #     Chrom_A=''
        # if udb_conv_cd:
        #     Cadmium_A=''
        # if udb_conv_eox:
        #     Eox_A='' 
        # if udb_conv_kw:
        #     Kw_A=''
        # if udb_conv_cu:
        #     Kupfer_A=''
        # if udb_conv_ni:
        #     Nickel_A=''
        # if udb_conv_pak16:
        #     Pak16_A=''
        # if udb_conv_pcb:
        #     Summe_PCB_A=''
        # if udb_conv_hg==True:
        #     Quecksilber_A=""
        # if udb_conv_zn:
        #     Zink_A=''
        
        
        with st.beta_expander("Feststoff"):
            st.markdown(
                '**Klassifikation innerhalb des Feststoffaliquotes**')
            st.write("Die Probe "+str(Probenbezeichnung_A) +
                     " wird gem. der bestimmten Feststoffparameter als "+zuordnungswert_sm_A+" klassifiziert.")
            if zuordnungswert_sm_A == "Z0":
                st.write("")
            elif zuordnungswert_sm_A == "Z1" or "Z1.1" or "Z1.2" or "Z2" or ">Z2":
                st.markdown(
                    "**Für die Feststoffe ist/sind der/die folgende/n Parameter für die Klassifikation als " + zuordnungswert_sm_A+" bestimmend**: "+str(bestimmende_Parameter_feststoff))
                st.write("\n")
            # PLOT #1 Feststoffe

            source = pd.DataFrame({'Parameter': ['Arsen', 'Blei', 'Cadmium', 'Chrom', 'Kupfer', 'Nickel', 'Quecksilber', 'Zink', 'EOX', 'KW', 'PAK16', 'PCB'],
                                   'Massenkonzentration': [Arsen_A, Blei_A, Cadmium_A, Chrom_A, Kupfer_A, Nickel_A, Quecksilber_A, Zink_A, Eox_A, Kw_A, Pak16_A, Summe_PCB_A],
                                   'udB': [udB_as, udB_pb, udB_cd, udB_cr, udB_cu, udB_ni, udB_hg, udB_zn, udB_eox, udB_kw, udB_pak16, udB_summepcb],
                                   'mg/kg': [O_Grenze_Z[0], O_Grenze_Z[1], O_Grenze_Z[2], O_Grenze_Z[3], O_Grenze_Z[4], O_Grenze_Z[5], O_Grenze_Z[6], O_Grenze_Z[7], O_Grenze_Z[8], O_Grenze_Z[9], O_Grenze_Z[10], O_Grenze_Z[11]],
                                   'Z': [Z_prior[0] + Z_next[0], Z_prior[1] + Z_next[1], Z_prior[2] + Z_next[2], Z_prior[3] + Z_next[3], Z_prior[4] + Z_next[4], Z_prior[5] + Z_next[5], Z_prior[6] + Z_next[6], Z_prior[7] + Z_next[7], Z_prior[8] + Z_next[8], Z_prior[9] + Z_next[9], Z_prior[10] + Z_next[10], Z_prior[11] + Z_next[11]],
                                   '%': [Arsen_A_erhöhung_plot, Blei_A_erhöhung_plot, Cadmium_A_erhöhung_plot, Chrom_A_erhöhung_plot, Kupfer_A_erhöhung_plot, Nickel_A_erhöhung_plot, Quecksilber_A_erhöhung_plot, Zink_A_erhöhung_plot, Eox_A_erhöhung_plot, Kw_A_erhöhung_plot, Pak16_A_erhöhung_plot, Summe_PCB_A_erhöhung_plot],
                                   'Z2': [Z_Rec_false_SM[0], Z_Rec_false_SM[1], Z_Rec_false_SM[2], Z_Rec_false_SM[3], Z_Rec_false_SM[4], Z_Rec_false_SM[5], Z_Rec_false_SM[6], Z_Rec_false_SM[7], Z_Rec_false_SM[8], Z_Rec_false_SM[9], Z_Rec_false_SM[10], Z_Rec_false_SM[11]]
                                   # 'As': [Arsen_A, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                                   })
            stepsize_feststoff_graph = 90
            bar = alt.Chart(source).mark_bar(color='#0040FF', opacity=0.3).encode(
                x='Parameter', y='Massenkonzentration').properties(width=alt.Step(stepsize_feststoff_graph))  # coole Farbe: #21B85B

            tick_höhe = -5

            tick = alt.Chart(source).mark_tick(color='blue', thickness=1, size=stepsize_feststoff_graph * 0.9).encode(x='Parameter', y='mg/kg').properties(
                title=Probenbezeichnung_A+" [Feststoff]                               Klassifikation [Feststoff]: "+zuordnungswert_sm_A)

            text2 = tick.mark_text(
                align='right',
                color='black',
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dx=-21,
                dy=tick_höhe
            ).encode(
                text='mg/kg'
            )
            tick2 = alt.Chart(source).mark_tick(
                text='green', thickness=0, size=stepsize_feststoff_graph * 0.9).encode(x='Parameter', y='mg/kg')
            text3 = tick2.mark_text(
                align='left',
                color='black',
                baseline='middle',
                fontSize=10,
                dy=-5,
                dx=-19,
            ).encode(
                text='Z'
            )
            tick3 = alt.Chart(source).mark_tick(
                text='green', thickness=0, size=stepsize_feststoff_graph * 0.9).encode(x='Parameter', y='mg/kg')
            text4 = tick3.mark_text(
                align='left',
                color=farbe_grenzlinie,
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dy=-15,
                dx=-10,
            ).encode(
                text='%'
            )

            text5 = tick3.mark_text(
                align='left',
                color=farbe_grenzlinie,
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dy=-35,
                dx=-10,
            ).encode(
                text='Z2'
            )
            text6 = tick3.mark_text(
                align='center',
                color='black',
                baseline='middle',
                fontStyle='bold',
                dy=-25,
                dx=0,
                # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                fontSize=12,
            ).encode(
                text='Massenkonzentration'
            )
            text7 = tick3.mark_text(
                align='center',
                color='black',
                baseline='middle',
                fontStyle='bold',
                dy=-25,
                dx=19,
                # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                fontSize=8.5,
            ).encode(
                text='udB'
            )
            st.altair_chart(bar+text6+tick+text2+tick2 +
                            text3+tick3+text4+text5+text7, use_container_width=False)

            Z_Klassifikation_eluat = z_el_A
            # print("ZELELUAT****")
            # print(z_el_A)
            # print(Z_Klassifikation_eluat)
            bestimmende_Parameter_liste_eluat = ['Arsen (Eluat) ', 'Blei (Eluat) ', 'Cadmium (Eluat) ', 'Chrom (Eluat) ',
                                                 'Kupfer (Eluat) ', 'Nickel (Eluat) ', 'Quecksilber (Eluat) ', 'Zink (Eluat) ', 'Chlorid ', 'Sulfat ', 'Phenol ', 'Leitfähigkeit']

            bestimmende_Parameter_eluat = []
            for i in range(0, len(Z_Klassifikation_eluat)) and range(0, len(bestimmende_Parameter_liste_eluat)):
                if float(max(Z_Klassifikation_eluat)) == Z_Klassifikation_eluat[i]:
                    bestimmender_parameter_eluat = bestimmende_Parameter_liste_eluat[i]
                    bestimmende_Parameter_eluat.append(
                        bestimmender_parameter_eluat)
                    # print(bestimmende_Parameter_eluat)
            for i in range(0, len(bestimmende_Parameter_eluat)):
                bestimmende_Parameter_eluat = "".join(
                    bestimmende_Parameter_eluat)

            if max(Z_Klassifikation_eluat) == 2.0 or 0.0:
                Z_Klassifikation_eluat = "Z"+str((max(Z_Klassifikation_eluat)))
            else:
                Z_Klassifikation_eluat = "Z" + \
                    str(float(max(Z_Klassifikation_eluat)))
            # print("Z-ELUAT")
            # print(Z_Klassifikation_eluat)
            if Z_Klassifikation_eluat == "Z0.0":
                Z_Klassifikation_eluat = "Z0"
                bem_eluat = "Bemerkungen: "
            elif Z_Klassifikation_eluat == "Z1.1":
                Z_Klassifikation_eluat = "Z1.1"
                bem_eluat = "Bemerkungen: " + \
                    str(bestimmende_Parameter_eluat)
            elif Z_Klassifikation_eluat == "Z1.2":
                Z_Klassifikation_eluat = "Z1.2"
                bem_eluat = "Bemerkungen: " + \
                    str(bestimmende_Parameter_eluat)
            elif Z_Klassifikation_eluat == "Z2.0" or "Z2":
                Z_Klassifikation_eluat = "Z2"
                bem_eluat = "Bemerkungen: " + \
                    str(bestimmende_Parameter_eluat)
            elif Z_Klassifikation_eluat == "Z9.0" or "Z9":
                Z_Klassifikation_eluat = ">Z2"
                bem_eluat = "Bemerkungen: " + \
                    str(bestimmende_Parameter_eluat)
            else:
                bem_eluat = "Bemerkungen: " + \
                    str(bestimmende_Parameter_eluat)
        #
        if pak_graph:
            PAK_Grenze = ["", "", "", "", "", "", "",
                          "", "", "", "", "", "", "", "", ""]
            with st.beta_expander("Verteilung der 16 Polyzyklischen Aromatischen Kohlenwasserstoffe der US-Environmental Protection Agency (16EPA-PAK)"):
                source = pd.DataFrame({'Parameter': ['Naphtalin', 'Acenaphtylen', 'Acenaphten', 'Fluoren', 'Phenanthren', 'Anthracen', 'Fluoranthen', 'Pyren', 'Benz(a)anthracen', 'Chrysen', 'Benzo(b)fluoranthen', 'Benzo(k)fluoranthen', 'Benzoapyren', 'Indeno(123-cd)pyren', 'Dibenz(ah)anthracen', 'Benzo(ghi)perylen'],
                                       'Massenkonzentration': [Naphthalin_A, Acenaphtylen_A, Acenaphten_A, Fluoren_A, Phenanthren_A, Anthracen_A, Fluoranthen_A, Pyren_A, Benzaanthracen_A, Chrysen_A, Benzobfluoranthen_A, Benzokfluoranthen_A, Benzoapyren_A, Indeno123cdpyren_A, Dibenzahanthracen_A, Benzoghiperylen_A],
                                       'udB': [udB_naphtalin, udB_acenaphtylen, udB_acenaphten, udB_fluoren, udB_Phenanthren, udB_Anthracen, udB_Fluoranthen, udB_Pyren, udB_Benzaanthracen, udB_Chrysen, udB_Benzobfluoranthen, udB_Benzokfluoranthen, udB_Benzoapyren, udB_Indeno123cdpyren, udB_Dibenzanthracen, udB_Benzoghiperylen],
                                       'mg/kg': [PAK_Grenze[0], PAK_Grenze[1], PAK_Grenze[2], PAK_Grenze[3], PAK_Grenze[4], PAK_Grenze[5], PAK_Grenze[6], PAK_Grenze[7], PAK_Grenze[8], PAK_Grenze[9], PAK_Grenze[10], PAK_Grenze[11], PAK_Grenze[12], PAK_Grenze[13], PAK_Grenze[14], PAK_Grenze[15]]
                                       })
                stepsize_pak_graph = 65
                bar_pak = alt.Chart(source).mark_bar(color='#0040FF', opacity=0.3).encode(
                    x='Parameter', y='Massenkonzentration').properties(width=alt.Step(stepsize_pak_graph))  # coole Farbe: #21B85B
                tick_pak = alt.Chart(source).mark_tick(
                    text='green', thickness=0, size=stepsize_pak_graph * 0.9).encode(x='Parameter', y='Massenkonzentration')
                text_pak_einheit = tick_pak.mark_text(
                    align='right',
                    color='black',
                    baseline='middle',
                    fontStyle='bold',
                    fontSize=10,
                    dx=-21,
                    dy=tick_höhe
                ).encode(
                    text='mg/kg'
                )
                tick_pak2 = alt.Chart(source).mark_tick(
                    text='green', thickness=0, size=stepsize_pak_graph * 0.9).encode(x='Parameter', y='Massenkonzentration')
                text_pak2 = tick_pak2.mark_text(
                    align='center',
                    color='black',
                    baseline='middle',
                    fontStyle='bold',
                    dy=-25,
                    dx=0,
                    # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                    fontSize=12,
                ).encode(
                    text='Massenkonzentration'
                )
                tick_pak3 = tick_pak.mark_text(
                    align='center',
                    color='black',
                    baseline='middle',
                    fontStyle='bold',
                    dy=-25,
                    dx=19,
                    # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                    fontSize=8.5,
                ).encode(
                    text='udB'
                )
                st.altair_chart(bar_pak+tick_pak+text_pak_einheit +
                                tick_pak2+text_pak2+tick_pak3, use_container_width=True)
                #st.markdown("**Angaben in mg/kg EINFÜGEN**")
        if pcb_graph:
            PCB_Grenze = ["", "", "", "", "", "", ""]
            with st.beta_expander("Verteilung der sechs Polychlorierten Biphenyle Kongenere (PCB) nach Balschmiter"):
                source = pd.DataFrame({'Parameter': ['PCB28', 'PCB52', 'PCB101', 'PCB138', 'PCB153', 'PCB180'],
                                       'Massenkonzentration': [Pcb28_A, Pcb52_A, Pcb101_A, Pcb138_A, Pcb153_A, Pcb180_A],
                                       'udB': [udBpcb28, udBpcb52, udBpcb101, udBpcb138, udBpcb153, udBpcb180],
                                       'mg/kg': [PCB_Grenze[0], PCB_Grenze[1], PCB_Grenze[2], PCB_Grenze[3], PCB_Grenze[4], PCB_Grenze[5]]
                                       })
                stepsize_pcb_graph = 65
                bar_pcb = alt.Chart(source).mark_bar(color='#0040FF', opacity=0.3).encode(
                    x='Parameter', y='Massenkonzentration').properties(width=alt.Step(stepsize_pcb_graph))  # coole Farbe: #21B85B
                tick_pcb = alt.Chart(source).mark_tick(
                    text='green', thickness=0, size=stepsize_pcb_graph * 0.9).encode(x='Parameter', y='Massenkonzentration')
                text_pcb_einheit = tick_pcb.mark_text(
                    align='right',
                    color='black',
                    baseline='middle',
                    fontStyle='bold',
                    fontSize=10,
                    dx=-21,
                    dy=tick_höhe
                ).encode(
                    text='mg/kg'
                )
                tick_pcb2 = alt.Chart(source).mark_tick(
                    text='green', thickness=0, size=stepsize_pcb_graph * 0.9).encode(x='Parameter', y='Massenkonzentration')
                text_pcb2 = tick_pcb2.mark_text(
                    align='center',
                    color='black',
                    baseline='middle',
                    fontStyle='bold',
                    dy=-25,
                    dx=0,
                    # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                    fontSize=12,
                ).encode(
                    text='Massenkonzentration'
                )
                tick_pcb3 = tick_pcb.mark_text(
                    align='center',
                    color='black',
                    baseline='middle',
                    fontStyle='bold',
                    dy=-25,
                    dx=19,
                    # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                    fontSize=8.5,
                ).encode(
                    text='udB'
                )
                st.altair_chart(bar_pcb+tick_pcb+text_pcb_einheit +
                                tick_pcb2+text_pcb2+tick_pcb3, use_container_width=True)
                #st.markdown("**Angaben in mg/kg EINFÜGEN**")

        with st.beta_expander("Eluat"):
            st.markdown(
                '**Dieser Abschnitt klassifiziert ausschließlich innerhalb des Eluates**')
            st.write("Die Probe "+str(Probenbezeichnung_A) +
                     " wird gem. der Analyse des Eluates als "+Z_Klassifikation_eluat+" klassifiziert.")
            if zuordnungswert_el_A == "Z0":
                st.write("")
            elif zuordnungswert_el_A == "Z1" or "Z1.1" or "Z1.2" or "Z2" or ">Z2":
                st.write(
                    "**Für die Eluate ist/sind der/die folgende/n Parameter als " + Z_Klassifikation_eluat + " bestimmend**: "+str(bestimmende_Parameter_eluat))
            # PLOT #1 Eluate
            # pH
            # print("ZREC ****")
            # print(Z_Rec_false_SM)
            #pH_carb = 13
            if pH_24 == True:

                # H=float(pH)
                if pH_depV_below == True and pH_depV==True:
                    st.error("Der pH-Wert ist nach LAGA mit "+str(pH_carb) +
                             " im sauren Milieu. Damit liegt der pH-Wert im DK-Bereich. Der pH-Wert entspricht dem pH-Wert nach Verringerung der Konzentration der Hydroniumionen durch Carbonatisierung durch CO2-Fixierung.")
                elif pH_depV == False and pH_depV_below==False:
                    st.success("Der pH-Wert ist nach LAGA mit "+str(pH_carb) +
                               " innerhalb des Toleranzbereiches von 7.0 bis 12.5.  Der pH-Wert entspricht dem pH-Wert nach Verringerung der Konzentration der Hydroniumionen durch Carbonatisierung durch CO2-Fixierung.")
                elif pH_depV_above == True and pH_depV==True:
                    st.error("Der pH-Wert ist nach LAGA mit "+str(pH_carb) +
                             " deutlich innerhalb des alkalischen Bereiches. Damit liegt der pH-Wert im DK-Bereich.  Der pH-Wert entspricht dem pH-Wert nach Verringerung der Konzentration der Hydroniumionen durch Carbonatisierung durch CO2-Fixierung.")
            elif pH_24 == False:
                # pH=float(pH)
                if pH_depV_below == True and pH_depV==True:
                    st.error("Der pH-Wert ist nach LAGA mit "+str(pH) +
                             " im sauren Milieu. Damit liegt der pH-Wert im DK-Bereich.")
                elif pH_depV == False and pH_depV_below==False:
                    st.success("Der pH-Wert ist nach LAGA mit "+str(pH) +
                               " innerhalb des Toleranzbereiches von 7.0 bis 12.5.")
                elif pH_depV_above == True and pH_depV==True:
                    st.error("Der pH-Wert ist nach LAGA mit "+str(pH) +
                             " deutlich innerhalb des alkalischen Bereiches. Damit liegt der pH-Wert im DK-Bereich.")
            else:
                st.error("Eine Angabe über den pH-Wert wurde nicht gefunden.")

            source = pd.DataFrame({'Parameter': ['Arsen (l)', 'Blei (l)', 'Cadmium (l)', 'Chrom (l)', 'Kupfer (l)', 'Nickel (l)', 'Quecksilber (l)', 'Zink (l)'],
                                   'Massenkonzentration': [As_el, Pb_el, Cd_el, Cr_el, Cu_el, Ni_el, Hg_el, Zn_el],
                                   'udB': [udB_as_el, udB_pb_el, udB_cd_el, udB_cr_el, udB_cu_el, udB_ni_el, udB_hg_el, udB_zn_el],
                                   'ug/L': [O_Grenze_Z[12], O_Grenze_Z[13], O_Grenze_Z[14], O_Grenze_Z[15], O_Grenze_Z[16], O_Grenze_Z[17], O_Grenze_Z[18], O_Grenze_Z[19]],
                                   'Z': [Z_prior[12] + Z_next[12], Z_prior[13] + Z_next[13], Z_prior[14] + Z_next[14], Z_prior[15] + Z_next[15], Z_prior[16] + Z_next[16], Z_prior[17] + Z_next[17], Z_prior[18] + Z_next[18], Z_prior[19] + Z_next[19]],
                                   '%': [As_el_erhöhung_plot, Pb_el_erhöhung_plot, Cd_el_erhöhung_plot, Cr_el_erhöhung_plot, Cu_el_erhöhung_plot, Ni_el_erhöhung_plot, Hg_el_erhöhung_plot, Zn_el_erhöhung_plot],
                                   'Z2': [Z_Rec_false_SM[12], Z_Rec_false_SM[13], Z_Rec_false_SM[14], Z_Rec_false_SM[15], Z_Rec_false_SM[16], Z_Rec_false_SM[17], Z_Rec_false_SM[18], Z_Rec_false_SM[19]]
                                   # 'As': [Arsen_A, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                                   })
            stepsize_eluat_graph = 90
            bar = alt.Chart(source).mark_bar(color='#0040FF', opacity=0.3).encode(
                x='Parameter', y='Massenkonzentration').properties(width=alt.Step(stepsize_eluat_graph))  # coole Farbe: #21B85B

            tick_höhe = -5

            tick = alt.Chart(source).mark_tick(color='blue', thickness=1, size=stepsize_eluat_graph * 0.9).encode(x='Parameter', y='ug/L').properties(
                title=Probenbezeichnung_A+" [Eluat]                               Klassifikation [Eluat]: " + Z_Klassifikation_eluat)  # zuordnungswert_el_A

            text2 = tick.mark_text(
                align='right',
                color='black',
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dx=-21,
                dy=tick_höhe
            ).encode(
                text='ug/L'
            )
            tick2 = alt.Chart(source).mark_tick(
                text='green', thickness=0, size=stepsize_eluat_graph * 0.9).encode(x='Parameter', y='ug/L')
            text3 = tick2.mark_text(
                align='left',
                color='black',
                baseline='middle',
                fontSize=10,
                dy=-5,
                dx=-19,
            ).encode(
                text='Z'
            )
            tick3 = alt.Chart(source).mark_tick(
                text='green', thickness=0, size=stepsize_eluat_graph * 0.9).encode(x='Parameter', y='ug/L')
            text4 = tick3.mark_text(
                align='left',
                color=farbe_grenzlinie,
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dy=-15,
                dx=-10,
            ).encode(
                text='%'
            )

            text5 = tick3.mark_text(
                align='left',
                color=farbe_grenzlinie,
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dy=-35,
                dx=-10,
            ).encode(
                text='Z2'
            )
            text6 = tick3.mark_text(
                align='center',
                color='black',
                baseline='middle',
                fontStyle='bold',
                dy=-25,
                dx=0,
                # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                fontSize=12,
            ).encode(
                text='Massenkonzentration'
            )
            text7 = tick3.mark_text(
                align='center',
                color='black',
                baseline='middle',
                fontStyle='bold',
                dy=-25,
                dx=19,
                # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                fontSize=8.5,
            ).encode(
                text='udB'
            )

            st.altair_chart(bar+text6+tick+text2+tick2 +
                            text3+tick3+text4+text5+text7, use_container_width=True)

            # PLOT #2 Eluate
            # print(Probenbezeichnung_A)
            source = pd.DataFrame({'Parameter': ['Chlorid (l)', 'Sulfat (l)', 'Phenol (l)', 'Leitfähigkeit'],
                                   'Massenkonzentration': [Cl_el, So4_el, Phenol, Leitfähigkeit],
                                   'mg/L': [O_Grenze_Z[20], O_Grenze_Z[21], O_Grenze_Z[22], O_Grenze_Z[23]],
                                   'Z': [Z_prior[20] + Z_next[20], Z_prior[21] + Z_next[21], Z_prior[22] + Z_next[22], Z_prior[23] + Z_next[23]],
                                   '%': [Cl_el_erhöhung_plot, So4_el_erhöhung_plot, Phenol_erhöhung_plot, Leitfähigkeit_erhöhung_plot],
                                   'Z2': [Z_Rec_false_SM[20], Z_Rec_false_SM[21], Z_Rec_false_SM[22], Z_Rec_false_SM[23]]
                                   # 'As': [Arsen_A, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                                   })
            stepsize_eluat_graph2 = 90
            bar = alt.Chart(source).mark_bar(color='#0040FF', opacity=0.3).encode(
                x='Parameter', y='Massenkonzentration').properties(width=alt.Step(stepsize_eluat_graph2))  # coole Farbe: #21B85B

            tick_höhe = -5

            tick = alt.Chart(source).mark_tick(color='blue', thickness=1, size=stepsize_eluat_graph2 * 0.9).encode(x='Parameter', y='mg/L').properties(
                title=Probenbezeichnung_A+" [Eluat]                               Klassifikation [Eluat]: " + Z_Klassifikation_eluat)  # zuordnungswert_el_A

            text2 = tick.mark_text(
                align='right',
                color='black',
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dx=-21,
                dy=tick_höhe
            ).encode(
                text='mg/L'
            )
            tick2 = alt.Chart(source).mark_tick(
                text='green', thickness=0, size=stepsize_eluat_graph2 * 0.9).encode(x='Parameter', y='mg/L')
            text3 = tick2.mark_text(
                align='left',
                color='black',
                baseline='middle',
                fontSize=10,
                dy=-5,
                dx=-19,
            ).encode(
                text='Z'
            )
            tick3 = alt.Chart(source).mark_tick(
                text='green', thickness=0, size=stepsize_eluat_graph2 * 0.9).encode(x='Parameter', y='mg/L')
            text4 = tick3.mark_text(
                align='left',
                color=farbe_grenzlinie,
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dy=-15,
                dx=-10,
            ).encode(
                text='%'
            )

            text5 = tick3.mark_text(
                align='left',
                color=farbe_grenzlinie,
                baseline='middle',
                fontStyle='bold',
                fontSize=10,
                dy=-35,
                dx=-10,
            ).encode(
                text='Z2'
            )
            text6 = tick3.mark_text(
                align='center',
                color='black',
                baseline='middle',
                fontStyle='bold',
                dy=-25,
                dx=0,
                # wurde ausgeschaltet, da analysierter Gehalt wo anders ausgegeben wird (text5)
                fontSize=12,
            ).encode(
                text='Massenkonzentration'
            )

            st.altair_chart(bar+text6+tick+text2+tick2 +
                            text3+tick3+text4+text5, use_container_width=True)

        if Recyclingbaustoffcheck_erweitert == False:
            st.error("Achtung: Es wurde gemäß der gewählten Einstellung keine .xlsx Datei mit den aus dem gewählten Prüfbericht extrahierten Parametern erzeugt. Die bestehenden .xlsx-Vorlagen knüpfen jedoch an die Option an, den Bauschutt als Recyclingbaustoff zu verwenden. ")
        else:

            if single_file:

                path = Path(
                    "C:/QLaga0.913 AlphaRelease/est2.xlsx")

            if '/' in Probenbezeichnung_A:
                Probenbezeichnung = Probenbezeichnung_A.replace(
                    "/", "_")
            elif '\n' in Probenbezeichnung_A:
                Probenbezeichnung = Probenbezeichnung_A.replace("\n", "")
                Probenbezeichnung = Probenbezeichnung.lstrip()
            else:
                Probenbezeichnung = Probenbezeichnung_A
            # C:/Users/0z/Desktop/Programmierung/Aktuelle Projekte/St0706/Streamlit
            path_est2 = r"C:/QLaga0.913 AlphaRelease/est2.xlsx"
            newpath1 = 'C:/QLaga0.913 AlphaRelease/'
            newpath2 = newpath1
            path_est3 = r"C:/QLaga0.913 AlphaRelease/est3.xlsx"
            # newpath2=
            # C:/Users/0z/Desktop/Programmierung/Aktuelle Projekte/St0706/Streamlit/
            print(os.path.getmtime(path_est2))
            print(os.path.getmtime(path_est3))
            time_excelausgabe1 = 1627375787.0
            errormessage_excel1 = True
            errormessage_excel2 = True
            if excelausgabe and os.path.getmtime(path_est2) == time_excelausgabe1:
                newpath = newpath1 + \
                    Probenbezeichnung+'.xlsx'
                if '\n' in newpath:
                    newpath = newpath.replace('\n', '')
                # print(newpath)
                copyfile(
                    path_est2, newpath)
                WorkBook = load_workbook(newpath, read_only=False)
                WorkSheet = WorkBook['Tabelle1']
                if udB_as == ' [u.d.B.]':
                    WorkSheet['E17'] = 'udB'
                else:
                    WorkSheet['E17'] = Arsen_A

                if udB_pb == ' [u.d.B.]':
                    WorkSheet['E20'] = 'udB'
                else:
                    WorkSheet['E20'] = Blei_A

                if udB_cd == ' [u.d.B.]':
                    WorkSheet['E19'] = 'udB'
                else:
                    WorkSheet['E19'] = Cadmium_A

                if udB_cr == ' [u.d.B.]':
                    WorkSheet['E21'] = 'udB'
                else:
                    WorkSheet['E21'] = Chrom_A

                if udB_cu == ' [u.d.B.]':
                    WorkSheet['E22'] = 'udB'
                else:
                    WorkSheet['E22'] = Kupfer_A

                if udB_ni == ' [u.d.B.]':
                    WorkSheet['E23'] = 'udB'
                else:
                    WorkSheet['E23'] = Nickel_A

                if udB_hg == ' [u.d.B.]':
                    WorkSheet['E18'] = 'udB'
                else:
                    WorkSheet['E18'] = Quecksilber_A

                if udB_zn == ' [u.d.B.]':
                    WorkSheet['E24'] = 'udB'
                else:
                    WorkSheet['E24'] = Zink_A

                if udB_eox == ' [u.d.B.]':
                    WorkSheet['E16'] = 'udB'
                else:
                    WorkSheet['E16'] = Eox_A

                if udB_summepcb == ' [u.d.B.]':
                    WorkSheet['E15'] = 'udB'
                else:
                    WorkSheet['E15'] = Summe_PCB_A

                if udB_kw == ' [u.d.B.]':
                    WorkSheet['E13'] = 'udB'
                else:
                    WorkSheet['E13'] = Kw_A

                WorkSheet['C3'] = Auftraggeber
                WorkSheet['C6'] = Probenbezeichnung_A
                WorkSheet['C8'] = Entnahmedatum_A
                WorkSheet['C55'] = zeitraumpruefung
                WorkSheet['C56'] = aktuellesDatum
                WorkSheet['E47'] = bem
                WorkSheet['C57'] = bearbeiter
                WorkSheet['G55'] = telefonnummer
                WorkSheet['G57'] = email
                WorkSheet['C4'] = Vorhabenbezeichnung
                WorkSheet['C5'] = Projektnummer
                if udB_as_el == ' [u.d.B.]':
                    WorkSheet['E35'] = 'udB'
                else:
                    WorkSheet['E35'] = As_el

                if udB_hg_el == ' [u.d.B.]':
                    WorkSheet['E36'] = 'udB'
                else:
                    WorkSheet['E36'] = Hg_el

                if udB_cd_el == ' [u.d.B.]':
                    WorkSheet['E37'] = 'udB'
                else:
                    WorkSheet['E37'] = Cd_el

                if udB_pb_el == ' [u.d.B.]':
                    WorkSheet['E38'] = 'udB'
                else:
                    WorkSheet['E38'] = Pb_el

                if udB_cr_el == ' [u.d.B.]':
                    WorkSheet['E39'] = 'udB'
                else:
                    WorkSheet['E39'] = Cr_el

                if udB_cu_el == ' [u.d.B.]':
                    WorkSheet['E40'] = 'udB'
                else:
                    WorkSheet['E40'] = Cu_el

                if udB_ni_el == ' [u.d.B.]':
                    WorkSheet['E41'] = 'udB'
                else:
                    WorkSheet['E41'] = Ni_el

                if udB_zn_el == ' [u.d.B.]':
                    WorkSheet['E42'] = 'udB'
                else:
                    WorkSheet['E42'] = Zn_el

                if udB_cl_el == ' [u.d.B.]':
                    WorkSheet['E33'] = 'udB'
                else:
                    WorkSheet['E33'] = Cl_el

                if udB_sulfat_el == ' [u.d.B.]':
                    WorkSheet['E34'] = 'udB'
                else:
                    WorkSheet['E34'] = So4_el

                if udB_pak16 == ' [u.d.B.]':
                    WorkSheet['E14'] = 'udB'
                else:
                    WorkSheet['E14'] = Pak16_A

                if udB_phenol_el == ' [u.d.B.]':
                    WorkSheet['E45'] = 'udB'
                else:
                    WorkSheet['E45'] = Phenol

                WorkSheet['E46'] = Z_Klassifikation
                WorkSheet['E32'] = pH_excel
                WorkSheet['E31'] = Leitfähigkeit

                WorkBook.save(newpath)
                excelausgabe_ = 1
            elif excelausgabe == True and errormessage_excel1 == True:
                st.error("Achtung: Die Excel-Datei wurde manipuliert oder ist fehlerhaft. Daher wurde keine erneute Datei erzeugt. Bitte die richtige Datei mit der richtigen Signatur nutzen.")
            
            timecode_excelausgabe_false = 1627464924.5803099

            if excelausgabe == False and os.path.getmtime(path_est3) == timecode_excelausgabe_false:

                # print("yeahhhhh")
                newpath = newpath2 + \
                    Probenbezeichnung+'.xlsx'
                if '\n' in newpath:
                    newpath = newpath.replace('\n', '')
                print(newpath)
                copyfile(path_est3, newpath)
                WorkBook = load_workbook(newpath, read_only=False)
                WorkSheet = WorkBook['Tabelle1']
                if udB_as == ' [u.d.B.]':
                    WorkSheet['E17'] = 'udB'
                else:
                    WorkSheet['E17'] = Arsen_A

                if udB_pb == ' [u.d.B.]':
                    WorkSheet['E20'] = 'udB'
                else:
                    WorkSheet['E20'] = Blei_A

                if udB_cd == ' [u.d.B.]':
                    WorkSheet['E19'] = 'udB'
                else:
                    WorkSheet['E19'] = Cadmium_A

                if udB_cr == ' [u.d.B.]':
                    WorkSheet['E21'] = 'udB'
                else:
                    WorkSheet['E21'] = Chrom_A

                if udB_cu == ' [u.d.B.]':
                    WorkSheet['E22'] = 'udB'
                else:
                    WorkSheet['E22'] = Kupfer_A

                if udB_ni == ' [u.d.B.]':
                    WorkSheet['E23'] = 'udB'
                else:
                    WorkSheet['E23'] = Nickel_A

                if udB_hg == ' [u.d.B.]':
                    WorkSheet['E18'] = 'udB'
                else:
                    WorkSheet['E18'] = Quecksilber_A

                if udB_zn == ' [u.d.B.]':
                    WorkSheet['E24'] = 'udB'
                else:
                    WorkSheet['E24'] = Zink_A

                if udB_eox == ' [u.d.B.]':
                    WorkSheet['E16'] = 'udB'
                else:
                    WorkSheet['E16'] = Eox_A

                if udB_summepcb == ' [u.d.B.]':
                    WorkSheet['E15'] = 'udB'
                else:
                    WorkSheet['E15'] = Summe_PCB_A

                if udB_kw == ' [u.d.B.]':
                    WorkSheet['E13'] = 'udB'
                    #WorkSheet['K13'] = 'udB'
                else:
                    WorkSheet['E13'] = Kw_A

                WorkSheet['C3'] = Auftraggeber
                WorkSheet['C6'] = Probenbezeichnung_A
                WorkSheet['C7'] = Entnahmedatum_A
                WorkSheet['C51'] = zeitraumpruefung
                WorkSheet['C52'] = aktuellesDatum
                WorkSheet['E43'] = bem
                WorkSheet['C53'] = bearbeiter
                WorkSheet['G51'] = telefonnummer
                WorkSheet['G52'] = email
                WorkSheet['C4'] = Vorhabenbezeichnung
                WorkSheet['C5'] = Projektnummer
                if udB_as_el == ' [u.d.B.]':
                    WorkSheet['E33'] = 'udB'
                else:
                    WorkSheet['E33'] = As_el

                if udB_hg_el == ' [u.d.B.]':
                    WorkSheet['E34'] = 'udB'
                else:
                    WorkSheet['E34'] = Hg_el

                if udB_cd_el == ' [u.d.B.]':
                    WorkSheet['E35'] = 'udB'
                else:
                    WorkSheet['E35'] = Cd_el

                if udB_pb_el == ' [u.d.B.]':
                    WorkSheet['E36'] = 'udB'
                else:
                    WorkSheet['E36'] = Pb_el

                if udB_cr_el == ' [u.d.B.]':
                    WorkSheet['E37'] = 'udB'
                else:
                    WorkSheet['E37'] = Cr_el

                if udB_cu_el == ' [u.d.B.]':
                    WorkSheet['E38'] = 'udB'
                else:
                    WorkSheet['E38'] = Cu_el

                if udB_ni_el == ' [u.d.B.]':
                    WorkSheet['E39'] = 'udB'
                else:
                    WorkSheet['E39'] = Ni_el

                if udB_zn_el == ' [u.d.B.]':
                    WorkSheet['E40'] = 'udB'
                else:
                    WorkSheet['E40'] = Zn_el

                if udB_cl_el == ' [u.d.B.]':
                    WorkSheet['E31'] = 'udB'
                else:
                    WorkSheet['E31'] = Cl_el

                if udB_sulfat_el == ' [u.d.B.]':
                    WorkSheet['E32'] = 'udB'
                else:
                    WorkSheet['E32'] = So4_el

                if udB_pak16 == ' [u.d.B.]':
                    WorkSheet['E14'] = 'udB'
                else:
                    WorkSheet['E14'] = Pak16_A

                if udB_phenol_el == ' [u.d.B.]':
                    WorkSheet['E41'] = 'udB'
                else:
                    WorkSheet['E41'] = Phenol

                WorkSheet['E42'] = Z_Klassifikation
                WorkSheet['E30'] = pH_excel
                WorkSheet['E29'] = Leitfähigkeit

                WorkBook.save(newpath)
                excelausgabe_ = 1
            elif errormessage_excel2 == True and excelausgabe == False:
                excelausgabe_manipuliert = 1
    if excelausgabe_ == 1 and excelausgabe == False:
        if Recyclingbaustoffcheck_erweitert == True or Recyclingbaustoffcheck_erweitert == False:
            st.success("Es wurde erfolgreich eine .xlsx-Datei (" +
                       Probenbezeichnung_A+") geschrieben, welche die Daten zusammenfasst.")
    if excelausgabe_manipuliert == 1:
        st.error("Achtung: Die Excel-Datei wurde manipuliert oder ist fehlerhaft. Daher wurde keine erneute Datei erzeugt. Bitte die richtige Datei mit der richtigen Signatur nutzen.")

elif active_tab == "LAGA Boden":
    st.subheader("Zuordnung nach LAGA Bauschutt M20 [2004]")
    with st.beta_expander("Optionen"):
        st.markdown('**Ausgewählte Option:**')
        option = st.radio("test", ('Bauschutt', 'Sand', 'Lehm/Schluff', 'Ton'))
        depVcheck = st.checkbox('DepV-Klassifikation ausgeben', value=True)
        st.markdown('**Ausgewählte Option:**')
        lagatext = '[LAGA (Juli 2004)](https://www.ngsmbh.de/bin/pdfs/Zuordnungswerte.pdf)'
        if option == 'Bauschutt' and depVcheck == True:
            st.write("Die Klassifikation für Bauschutt erfolgt nach "+lagatext +
                     ". Die Klassifikationen der Einzelparameter nach DepV wird zusätzlich ausgegeben.")
            # st.markdown(erlass, unsafe_allow_html=True) # hier klären wieso das nicht geht !
        elif option == 'Bauschutt' and depVcheck == False:
            st.write("Die Klassifikation für Bauschutt erfolgt nach "+lagatext+".")
        elif option == 'Sand' and depVcheck == False:
            st.write("Die Klassifikation für Sand erfolgt nach "+lagatext+".")
            st.error('noch nicht implementiert')
        elif option == 'Sand' and depVcheck == True:
            st.write("Die Klassifikation für Sand erfolgt nach "+lagatext +
                     ". Die Klassifikationen der Einzelparameter nach DepV wird zusätzlich ausgegeben.")
            st.error('noch nicht implementiert')
        elif option == 'Lehm/Schluff' and depVcheck == False:
            st.write(
                "Die Klassifikation für Lehm/Schluff erfolgt nach "+lagatext+".")
            st.error('noch nicht implementiert')
        elif option == 'Lehm/Schluff' and depVcheck == True:
            st.write("Die Klassifikation für Lehm/Schluff erfolgt nach "+lagatext +
                     ".Die Klassifikationen der Einzelparameter nach DepV wird zusätzlich ausgegeben.")
            st.error('noch nicht implementiert')
        elif option == 'Ton' and depVcheck == False:
            st.write("Die Klassifikation für Ton erfolgt nach "+lagatext+".")
            st.error('noch nicht implementiert')
        elif option == 'Ton' and depVcheck == True:
            st.write("Die Klassifikation für Ton erfolgt nach "+lagatext +
                     ".Die Klassifikationen der Einzelparameter nach DepV wird zusätzlich ausgegeben.")
            st.error('noch nicht implementiert')
        else:
            st.write('Es erfolgt eine Einstufung nach '+lagatext +
                     '. Die DepV (2009) wird zzgl. des '+erlass+" berücksichtigt.")
            check_depV = 1

elif active_tab == "Gebäude-Wiki":
    st.success("Hier kommt ein Wiki hin")

elif active_tab == "Geotechnik":
    st.success("Auswertehilfen Lastplattendruckversuch")

elif active_tab == "Hydrogeologie":
    st.success("Auswertehilfen Hydrogeologie")

elif active_tab == "Anregungen und Hilfe":
    st.write("Wendet euch bei Fragen am besten an:")
    st.write("T. Frank \n\n thomas.frank@mail.de")
    st.success('Literatur-Schlüssel: [Author Jahr] Titel, Datum')
    st.success(
        'Beispiel: [Grunert 2003] Auswertung des dynamischen Lastplattendruckversuches, 24.05.2003')

else:
    st.error("Something has gone terribly wrong.")


